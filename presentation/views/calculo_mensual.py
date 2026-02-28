import json
import calendar
import streamlit as st
import pandas as pd
import io
from datetime import datetime

# Librerías para Excel Corporativo
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Librerías para PDF Corporativo
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, portrait, letter, legal
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# Base de Datos Neon
from sqlalchemy import or_
from infrastructure.database.connection import SessionLocal
from infrastructure.database.models import Trabajador, Concepto, ParametroLegal, VariablesMes, PlanillaMensual, Prestamo, CuotaPrestamo


# ─── HELPERS DE BASE DE DATOS ───────────────────────────────────────────────

def _cargar_parametros(db, empresa_id, periodo_key) -> dict | None:
    """Lee los ParametroLegal de Neon y los devuelve como dict con claves compatibles."""
    p_db = db.query(ParametroLegal).filter_by(
        empresa_id=empresa_id, periodo_key=periodo_key
    ).first()
    if not p_db:
        return None
    return {
        'rmv': p_db.rmv, 'uit': p_db.uit,
        'tasa_onp': p_db.tasa_onp, 'tasa_essalud': p_db.tasa_essalud,
        'tasa_eps': p_db.tasa_eps, 'tope_afp': p_db.tope_afp,
        'afp_habitat_aporte': p_db.h_ap, 'afp_habitat_prima': p_db.h_pr,
        'afp_habitat_flujo': p_db.h_fl, 'afp_habitat_mixta': p_db.h_mx,
        'afp_integra_aporte': p_db.i_ap, 'afp_integra_prima': p_db.i_pr,
        'afp_integra_flujo': p_db.i_fl, 'afp_integra_mixta': p_db.i_mx,
        'afp_prima_aporte': p_db.p_ap, 'afp_prima_prima': p_db.p_pr,
        'afp_prima_flujo': p_db.p_fl, 'afp_prima_mixta': p_db.p_mx,
        'afp_profuturo_aporte': p_db.pr_ap, 'afp_profuturo_prima': p_db.pr_pr,
        'afp_profuturo_flujo': p_db.pr_fl, 'afp_profuturo_mixta': p_db.pr_mx,
        'tasa_4ta': getattr(p_db, 'tasa_4ta', 8.0) or 8.0,
        'tope_4ta': getattr(p_db, 'tope_4ta', 1500.0) or 1500.0,
    }


def _cargar_trabajadores_df(db, empresa_id) -> pd.DataFrame:
    """Lee trabajadores activos de Neon y los devuelve como DataFrame compatible."""
    trabajadores = (
        db.query(Trabajador)
        .filter_by(empresa_id=empresa_id, situacion="ACTIVO")
        .filter(or_(
            Trabajador.tipo_contrato == 'PLANILLA', 
            Trabajador.tipo_contrato == None,
            Trabajador.tipo_contrato == ''
        ))
        .all()
    )
    rows = []
    for t in trabajadores:
        rows.append({
            "Num. Doc.": t.num_doc,
            "Nombres y Apellidos": t.nombres,
            "Fecha Ingreso": t.fecha_ingreso,
            "Sueldo Base": t.sueldo_base,
            "Sistema Pensión": t.sistema_pension or "NO AFECTO",
            "Comisión AFP": t.comision_afp or "FLUJO",
            "Asig. Fam.": "Sí" if t.asig_fam else "No",
            "EPS": "Sí" if t.eps else "No",
            "CUSPP": t.cuspp or "",
            "Cargo": t.cargo or "",
            "Seguro Social": getattr(t, 'seguro_social', None) or "ESSALUD",
            "Banco": t.banco or "",
            "Cuenta Bancaria": t.cuenta_bancaria or "",
            "CCI": t.cci or "",
        })
    return pd.DataFrame(rows)


def _cargar_variables_df(db, empresa_id, periodo_key, conceptos) -> pd.DataFrame:
    """Lee VariablesMes de Neon y los devuelve como DataFrame compatible."""
    variables = (
        db.query(VariablesMes)
        .filter_by(empresa_id=empresa_id, periodo_key=periodo_key)
        .all()
    )
    concepto_nombres = [c.nombre for c in conceptos]
    rows = []
    for v in variables:
        susp = json.loads(getattr(v, 'suspensiones_json', '{}') or '{}')
        # Total de ausencias desde suspensiones_json; fallback a dias_faltados
        total_ausencias = sum(susp.values()) if susp else (v.dias_faltados or 0)
        row = {
            "Num. Doc.": v.trabajador.num_doc,
            "Nombres y Apellidos": v.trabajador.nombres,
            "Días Faltados": total_ausencias,
            "suspensiones_json": json.dumps(susp),
            "Min. Tardanza": v.min_tardanza or 0,
            "Hrs Extras 25%": v.hrs_extras_25 or 0.0,
            "Hrs Extras 35%": v.hrs_extras_35 or 0.0,
        }
        conceptos_data = json.loads(v.conceptos_json or '{}')
        for nombre in concepto_nombres:
            row[nombre] = conceptos_data.get(nombre, 0.0)
        rows.append(row)
    return pd.DataFrame(rows).fillna(0.0)


def _cargar_conceptos_df(db, empresa_id) -> pd.DataFrame:
    """Lee conceptos de la empresa de Neon como DataFrame compatible con el motor."""
    conceptos = db.query(Concepto).filter_by(empresa_id=empresa_id).all()
    rows = []
    for c in conceptos:
        rows.append({
            "Empresa_ID": empresa_id,
            "Nombre del Concepto": c.nombre,
            "Tipo": c.tipo,
            "Afecto AFP/ONP": c.afecto_afp,
            "Afecto 5ta Cat.": c.afecto_5ta,
            "Afecto EsSalud": c.afecto_essalud,
            "Computable CTS": c.computable_cts,
            "Computable Grati": c.computable_grati,
            "Prorrateable": getattr(c, 'prorrateable_por_asistencia', False),
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _guardar_planilla(db, empresa_id, periodo_key, df_resultados, auditoria_data):
    """Guarda (upsert) el resultado de planilla en la tabla PlanillaMensual de Neon."""
    resultado_json = df_resultados.to_json(orient='records', date_format='iso')
    auditoria_json = json.dumps(auditoria_data, default=str)

    existente = db.query(PlanillaMensual).filter_by(
        empresa_id=empresa_id, periodo_key=periodo_key
    ).first()
    if existente:
        existente.resultado_json = resultado_json
        existente.auditoria_json = auditoria_json
        existente.fecha_calculo = datetime.now()
    else:
        nueva = PlanillaMensual(
            empresa_id=empresa_id,
            periodo_key=periodo_key,
            resultado_json=resultado_json,
            auditoria_json=auditoria_json,
        )
        db.add(nueva)
    db.commit()


def _cargar_planilla_guardada(db, empresa_id, periodo_key):
    """Recupera una planilla previamente guardada de Neon. Retorna (df, auditoria) o (None, None)."""
    p = db.query(PlanillaMensual).filter_by(
        empresa_id=empresa_id, periodo_key=periodo_key
    ).first()
    if not p:
        return None, None
    df = pd.read_json(io.StringIO(p.resultado_json), orient='records')
    auditoria = json.loads(p.auditoria_json)
    return df, auditoria

MESES = ["01 - Enero", "02 - Febrero", "03 - Marzo", "04 - Abril", "05 - Mayo", "06 - Junio", 
         "07 - Julio", "08 - Agosto", "09 - Septiembre", "10 - Octubre", "11 - Noviembre", "12 - Diciembre"]

# --- 1. GENERADORES DE EXPORTACIÓN ---

_MESES_ES_CALC = {
    "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
    "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
}

def _periodo_legible_calc(periodo_key: str) -> str:
    """'02-2026' → 'Febrero - 2026'"""
    partes = periodo_key.split("-")
    if len(partes) == 2:
        return f"{_MESES_ES_CALC.get(partes[0], partes[0])} - {partes[1]}"
    return periodo_key

# Columnas internas que NO deben aparecer en los reportes (aliases/duplicados)
_COLS_OCULTAS_SABANA = {"EsSalud Patronal"}


def generar_excel_sabana(df, empresa_nombre, periodo, empresa_ruc=""):
    """Genera un archivo Excel nativo (.xlsx) profesional con colores corporativos"""
    periodo_texto = _periodo_legible_calc(periodo)
    # Excluir columna duplicada
    cols_mostrar = [c for c in df.columns if c not in _COLS_OCULTAS_SABANA]
    df_export = df[cols_mostrar]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name=f'Planilla_{periodo[:2]}', index=False, startrow=5)
        ws = writer.sheets[f'Planilla_{periodo[:2]}']

        # 1. Títulos de Cabecera
        ws['A1'] = empresa_nombre
        ws['A1'].font = Font(size=16, bold=True, color="0F2744")
        ws['A2'] = empresa_ruc and f"RUC: {empresa_ruc}" or ""
        ws['A2'].font = Font(size=10, color="64748B")
        ws['A3'] = f"PLANILLA DE REMUNERACIONES — PERIODO: {periodo_texto}"
        ws['A3'].font = Font(size=11, bold=True, color="1E4D8C")
        ws['A4'] = f"Fecha de Cálculo: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A4'].font = Font(size=10, italic=True, color="7F8C8D")
        
        # 2. Estilos Corporativos
        fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        
        # 3. Aplicar formato a las celdas de la tabla (datos desde fila 6)
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_thin
                if cell.row == 6:  # Fila de nombres de columnas
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = align_center
                elif cell.row == ws.max_row:  # Última fila (Totales)
                    cell.fill = fill_total
                    cell.font = Font(bold=True)
                    
        # 4. Auto-ajustar ancho de columnas (máximo 25 caracteres para no hacerlo gigante)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 25)
            
    buffer.seek(0)
    return buffer

def generar_pdf_sabana(df, empresa_nombre, periodo, empresa_ruc="", empresa_regimen=""):
    """Genera la sábana principal de planilla — ajustada a hoja, sin overflow."""
    periodo_texto = _periodo_legible_calc(periodo)

    # 1. Excluir columnas internas/duplicadas y datos bancarios (exclusivos del Reporte Tesorería)
    _OCULTAR_PDF = _COLS_OCULTAS_SABANA | {"Banco", "N° Cuenta", "CCI"}
    cols_work = [c for c in df.columns if c not in _OCULTAR_PDF]
    df = df[cols_work].copy()

    # 2. Excluir columnas numéricas donde TODOS los trabajadores tienen 0
    #    (ej. "Otros Ingresos" si nadie tiene bonos extra). Se compara solo sobre filas de datos, no la fila TOTALES.
    _COLS_IDENTIDAD = {"N°", "DNI", "Apellidos y Nombres", "Sist. Pensión", "Seg. Social"}
    if 'Apellidos y Nombres' in df.columns:
        df_datos = df[df['Apellidos y Nombres'] != 'TOTALES']
    else:
        df_datos = df.iloc[:-1]
    cols_mostrar = [
        c for c in df.columns
        if c in _COLS_IDENTIDAD
        or (c not in df_datos.columns)
        or pd.to_numeric(df_datos[c], errors='coerce').fillna(0).sum() != 0
    ]
    df = df[cols_mostrar]

    # ── Página landscape legal (1008 × 612 pt), márgenes 12 ──
    W_PAGE = 1008 - 24   # 984 pt disponibles
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=15, bottomMargin=15
    )
    elements = []
    styles = getSampleStyleSheet()

    C_NAVY  = colors.HexColor("#0F2744")
    C_STEEL = colors.HexColor("#1E4D8C")
    C_GOLD  = colors.HexColor("#C9A84C")
    C_LIGHT = colors.HexColor("#F0F4F9")
    C_GRAY  = colors.HexColor("#64748B")

    st_title = ParagraphStyle('T', fontName="Helvetica-Bold", fontSize=14, textColor=C_NAVY, spaceAfter=6)
    st_sub   = ParagraphStyle('S', fontName="Helvetica",      fontSize=9,  textColor=C_GRAY, spaceAfter=1)
    st_head  = ParagraphStyle('H', fontName="Helvetica-Bold", fontSize=10, textColor=C_STEEL, spaceAfter=8, spaceBefore=4)

    fecha_calc = datetime.now().strftime("%d/%m/%Y %H:%M")
    ruc_line = f"  |  RUC: {empresa_ruc}" if empresa_ruc else ""
    reg_line = f"  |  {empresa_regimen}" if empresa_regimen else ""
    elements.append(Paragraph(empresa_nombre + ruc_line, st_title))
    elements.append(Paragraph(
        f"PLANILLA DE REMUNERACIONES  ·  PERIODO: {periodo_texto}{reg_line}", st_head
    ))
    elements.append(Paragraph(f"Fecha de cálculo: {fecha_calc}", st_sub))
    elements.append(Spacer(1, 8))

    # ── Anchos de columna proporcionales (suman W_PAGE) ──
    # Orden de columnas esperado tras excluir EsSalud Patronal:
    col_widths_map = {
        "N°": 18, "DNI": 52, "Apellidos y Nombres": 108,
        "Sist. Pensión": 58, "Seg. Social": 50,
        "Sueldo Base": 52, "Asig. Fam.": 38, "Otros Ingresos": 52,
        "TOTAL BRUTO": 52, "ONP (13%)": 44, "AFP Aporte": 44,
        "AFP Seguro": 44, "AFP Comis.": 44, "Ret. 5ta Cat.": 46,
        "Dsctos/Faltas": 46, "NETO A PAGAR": 55, "Aporte Seg. Social": 58,
    }
    col_w = [col_widths_map.get(c, 48) for c in cols_mostrar]
    # Escalar para ocupar exactamente W_PAGE
    total_w = sum(col_w)
    col_w = [w * W_PAGE / total_w for w in col_w]

    # ── Construir datos de la tabla ──
    # Cabeceras más cortas para encabezado (wrapeadas con Paragraph)
    hdr_style = ParagraphStyle('HDR', fontName="Helvetica-Bold", fontSize=6.5,
                               textColor=colors.white, alignment=1, leading=8)
    p_nom_style = ParagraphStyle('NOM', fontName="Helvetica", fontSize=6.5,
                                 textColor=colors.black, alignment=0, leading=8, wordWrap='LTR')
    # Índice de la columna "Apellidos y Nombres" para aplicar word-wrap
    nom_idx = cols_mostrar.index("Apellidos y Nombres") if "Apellidos y Nombres" in cols_mostrar else -1
    data_rows = [[Paragraph(c, hdr_style) for c in cols_mostrar]]

    for _, row in df.iterrows():
        fila = []
        for i, val in enumerate(row):
            if i == nom_idx:
                fila.append(Paragraph(str(val) if str(val) != "nan" else "", p_nom_style))
            elif isinstance(val, float):
                fila.append(f"{val:,.2f}")
            else:
                fila.append(str(val) if str(val) != "nan" else "")
        data_rows.append(fila)

    t = Table(data_rows, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  C_STEEL),
        ('TEXTCOLOR',     (0,0),  (-1,0),  colors.white),
        ('ALIGN',         (0,0),  (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (2,1),  (2,-1),  'LEFT'),    # Nombres a la izquierda
        ('FONTNAME',      (0,1),  (-1,-2), 'Helvetica'),
        ('FONTSIZE',      (0,1),  (-1,-2), 6.5),
        ('TOPPADDING',    (0,0),  (-1,-1), 4),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 4),
        ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, C_LIGHT]),
        ('BACKGROUND',    (0,-1), (-1,-1), colors.HexColor("#CBD5E1")),
        ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,-1), (-1,-1), 6.5),
        ('GRID',          (0,0),  (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ('LINEABOVE',     (0,-1), (-1,-1), 0.8, C_NAVY),
        ('LINEBELOW',     (0,0),  (-1,0),  0.8, C_GOLD),
    ]))
    elements.append(t)
    
    # --- CUADRO RESUMEN DE PREVISIONES (AFP/ONP/SEGURIDAD SOCIAL/5TA) ---
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("<b>RESUMEN GENERAL DE OBLIGACIONES DEL EMPLEADOR Y RETENCIONES</b>", st_sub))
    elements.append(Spacer(1, 10))

    df_data = df.iloc[:-1]  # Sin fila de totales
    resumen_data = [["CONCEPTO", "DETALLE", "MONTO (S/)  "]]
    total_general = 0.0

    # ONP
    if 'ONP (13%)' in df_data.columns:
        onp_total = df_data['ONP (13%)'].sum()
        if onp_total > 0:
            n_onp = len(df_data[df_data['ONP (13%)'] > 0])
            resumen_data.append(["RETENCIÓN ONP (13%)", f"{n_onp} trabajador(es)", f"{onp_total:,.2f}"])
            total_general += onp_total

    # AFP por entidad — columnas pueden no existir si el filtro de ceros las eliminó
    if 'Sist. Pensión' in df_data.columns:
        for afp in df_data['Sist. Pensión'].unique():
            if "AFP" in str(afp):
                df_afp = df_data[df_data['Sist. Pensión'] == afp]
                tot = (
                    (df_afp['AFP Aporte'].sum() if 'AFP Aporte' in df_afp.columns else 0.0) +
                    (df_afp['AFP Seguro'].sum() if 'AFP Seguro' in df_afp.columns else 0.0) +
                    (df_afp['AFP Comis.'].sum() if 'AFP Comis.' in df_afp.columns else 0.0)
                )
                if tot > 0:
                    resumen_data.append([f"RETENCIÓN {afp}", f"{len(df_afp)} trabajador(es)", f"{tot:,.2f}"])
                    total_general += tot

    # Renta 5ta Categoría
    if 'Ret. 5ta Cat.' in df_data.columns:
        quinta_total = df_data['Ret. 5ta Cat.'].sum()
        if quinta_total > 0:
            n_quinta = len(df_data[df_data['Ret. 5ta Cat.'] > 0])
            resumen_data.append(["RETENCIÓN RENTA 5TA CAT.", f"{n_quinta} trabajador(es)", f"{quinta_total:,.2f}"])
            total_general += quinta_total

    resumen_data.append(["TOTAL RETENCIONES (A DECLARAR PDT)", "", f"S/ {total_general:,.2f}"])

    # Seguridad Social (ESSALUD + SIS) — a cargo del empleador
    resumen_data.append(["", "", ""])
    aporte_seg_total = 0.0
    col_seg = 'Aporte Seg. Social' if 'Aporte Seg. Social' in df_data.columns else 'EsSalud Patronal'
    if col_seg in df_data.columns and 'Seg. Social' in df_data.columns:
        for tipo_seg in df_data['Seg. Social'].unique():
            df_seg = df_data[df_data['Seg. Social'] == tipo_seg]
            monto_seg = df_seg[col_seg].sum()
            if monto_seg > 0:
                resumen_data.append([f"APORTE {tipo_seg} (EMPLEADOR)", f"{len(df_seg)} trabajador(es)", f"{monto_seg:,.2f}"])
                aporte_seg_total += monto_seg
    elif col_seg in df_data.columns:
        aporte_seg_total = df_data[col_seg].sum()
        resumen_data.append(["APORTE ESSALUD (EMPLEADOR)", f"{len(df_data)} trabajador(es)", f"{aporte_seg_total:,.2f}"])

    if aporte_seg_total > 0:
        resumen_data.append(["TOTAL SEGURIDAD SOCIAL (A PAGAR SUNAT)", "", f"S/ {aporte_seg_total:,.2f}"])

    t_res = Table(resumen_data, colWidths=[250, 150, 150])
    estilos_res = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#34495E")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#BDC3C7")),
    ]
    # Filas de subtotal en negrita con fondo
    for i, row_data in enumerate(resumen_data):
        if row_data[0].startswith("TOTAL"):
            estilos_res.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#E5E7E9")))
            estilos_res.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
    t_res.setStyle(TableStyle(estilos_res))

    if len(resumen_data) > 2:
        elements.append(t_res)

    doc.build(elements)
    buffer.seek(0)
    return buffer

def generar_pdf_quinta(data_q, empresa_nombre, periodo, nombre_trabajador):
    """Genera el Certificado Corporativo de Retención de 5ta Categoría"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=portrait(letter), rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, textColor=colors.HexColor("#2C3E50"), spaceAfter=5)
    subtitle_style = ParagraphStyle('CustomSub', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor("#7F8C8D"), alignment=TA_CENTER, spaceAfter=20)
    header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=10, textColor=colors.black, spaceAfter=5)

    elements.append(Paragraph(f"<b>{empresa_nombre}</b>", title_style))
    elements.append(Paragraph(f"LIQUIDACIÓN Y DETALLE DE RETENCIÓN DE 5TA CATEGORÍA", subtitle_style))
    elements.append(Paragraph(f"<b>Trabajador:</b> {nombre_trabajador}", header_style))
    elements.append(Paragraph(f"<b>Periodo de Cálculo:</b> {periodo}", header_style))
    elements.append(Spacer(1, 15))

    datos_tabla = [
        ["CONCEPTO / PASO DE CÁLCULO (SUNAT)", "IMPORTES (S/)"],
        ["1. Remuneraciones Previas Percibidas (Ene - Mes Anterior)", f"{int(data_q['rem_previa']):,}"],
        ["2. Remuneración Computable del Mes Actual", f"{data_q['base_mes']:,.2f}"],
        [f"3. Proyección de Meses Restantes ({data_q['meses_restantes']} meses)", f"{data_q['proy_sueldo']:,.2f}"],
        ["4. Proyección de Gratificaciones + Bono Extraordinario 9%", f"{data_q['proy_grati']:,.2f}"],
        ["A. RENTA BRUTA ANUAL PROYECTADA (REDONDEADA)", f"{int(data_q['bruta_anual']):,}"],
        [f"B. Deducción de Ley (7 UIT de S/ {data_q['uit_valor']:,.2f})", f"- {int(data_q['uit_7']):,}"],
        ["C. RENTA NETA IMPONIBLE ANUAL (A - B)", f"{int(data_q['neta_anual']):,}"]
    ]

    t_calc = Table(datos_tabla, colWidths=[380, 100])
    t_calc.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1f77b4")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,5), (-1,5), 'Helvetica-Bold'), 
        ('FONTNAME', (0,7), (-1,7), 'Helvetica-Bold'), 
        ('BACKGROUND', (0,7), (-1,7), colors.HexColor("#ECF0F1")),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7"))
    ]))
    elements.append(t_calc)
    elements.append(Spacer(1, 15))

    if data_q['neta_anual'] > 0:
        elements.append(Paragraph("<b>D. APLICACIÓN DE ESCALAS Y TRAMOS DEL IMPUESTO:</b>", header_style))
        elements.append(Spacer(1, 5))
        
        datos_tramos = [["TRAMO (UIT)", "TASA", "BASE APLICABLE", "IMPUESTO ANUAL"]]
        for tramo in data_q['detalle_tramos']:
            datos_tramos.append([tramo['rango'], tramo['tasa'], f"{int(tramo['base']):,}", f"{int(tramo['impuesto']):,}"])
        
        datos_tramos.append(["", "", "IMPUESTO ANUAL TOTAL:", f"{int(data_q['imp_anual']):,}"])
        
        t_tramos = Table(datos_tramos, colWidths=[150, 60, 130, 140])
        t_tramos.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#7F8C8D")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7"))
        ]))
        elements.append(t_tramos)
        elements.append(Spacer(1, 15))

        datos_final = [
            ["E. Impuesto Anual Calculado", f"{int(data_q['imp_anual']):,}"],
            ["F. Retenciones de Meses Anteriores Efectuadas", f"- {int(data_q['ret_previa']):,}"],
            [f"G. Divisor Aplicable al Mes ({data_q['divisor']})", f"÷ {data_q['divisor']}"],
            ["H. RETENCIÓN EXACTA A EFECTUAR EN EL MES (REDONDEADA)", f"S/ {int(data_q['retencion']):,}"]
        ]
        t_final = Table(datos_final, colWidths=[380, 100])
        t_final.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#F9E79F")), 
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7"))
        ]))
        elements.append(t_final)
    else:
        elements.append(Paragraph("<i>El trabajador no supera las 7 UIT, por lo tanto, la Renta Neta es S/ 0.00 y no aplica retención en este periodo.</i>", header_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer
def generar_excel_honorarios(df_loc, empresa_nombre, periodo_key, empresa_ruc=""):
    """Genera Excel corporativo para valorización de locadores."""
    periodo_texto = _periodo_legible_calc(periodo_key)
    # Limpiar columnas no deseadas para el reporte corporativo
    cols_excluir = ["Banco", "N° Cuenta", "CCI", "Observaciones"]
    df_export = df_loc[[c for c in df_loc.columns if c not in cols_excluir]].copy()
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name=f'Honorarios_{periodo_key[:2]}', index=False, startrow=5)
        ws = writer.sheets[f'Honorarios_{periodo_key[:2]}']
        ws['A1'] = empresa_nombre
        ws['A1'].font = Font(size=16, bold=True, color="0F2744")
        ws['A2'] = f"RUC: {empresa_ruc}" if empresa_ruc else ""
        ws['A2'].font = Font(size=10, color="64748B")
        ws['A3'] = f"VALORIZACIÓN DE LOCADORES DE SERVICIO (4ta Categoría) — PERIODO: {periodo_texto}"
        ws['A3'].font = Font(size=11, bold=True, color="1E4D8C")
        ws['A4'] = f"Fecha de Cálculo: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A4'].font = Font(size=10, italic=True, color="7F8C8D")
        fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_thin
                if cell.row == 6:
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = align_center
                elif cell.row == ws.max_row:
                    cell.fill = fill_total
                    cell.font = Font(bold=True)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 25)
    buffer.seek(0)
    return buffer


def generar_pdf_honorarios(df_loc, empresa_nombre, periodo_key, empresa_ruc="", empresa_regimen=""):
    """Genera PDF corporativo para valorización de locadores (igual diseño que sábana planilla)."""
    periodo_texto = _periodo_legible_calc(periodo_key)
    W_PAGE = 1008 - 24
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=15, bottomMargin=15
    )
    elements = []
    C_NAVY  = colors.HexColor("#0F2744")
    C_STEEL = colors.HexColor("#1E4D8C")
    C_GOLD  = colors.HexColor("#C9A84C")
    C_LIGHT = colors.HexColor("#F0F4F9")
    C_GRAY  = colors.HexColor("#64748B")
    st_title = ParagraphStyle('T', fontName="Helvetica-Bold", fontSize=14, textColor=C_NAVY, spaceAfter=6)
    st_sub   = ParagraphStyle('S', fontName="Helvetica",      fontSize=9,  textColor=C_GRAY, spaceAfter=1)
    st_head  = ParagraphStyle('H', fontName="Helvetica-Bold", fontSize=10, textColor=C_STEEL, spaceAfter=8, spaceBefore=4)
    fecha_calc = datetime.now().strftime("%d/%m/%Y %H:%M")
    ruc_line = f"  |  RUC: {empresa_ruc}" if empresa_ruc else ""
    reg_line = f"  |  {empresa_regimen}" if empresa_regimen else ""
    elements.append(Paragraph(empresa_nombre + ruc_line, st_title))
    elements.append(Paragraph(
        f"VALORIZACIÓN DE LOCADORES DE SERVICIO (4ta CATEGORÍA)  ·  PERIODO: {periodo_texto}{reg_line}", st_head
    ))
    elements.append(Paragraph(f"Fecha de cálculo: {fecha_calc}", st_sub))
    elements.append(Spacer(1, 8))

    # Limpiar columnas para PDF corporativo
    cols_excluir = ["Banco", "N° Cuenta", "CCI", "Observaciones"]
    cols = [c for c in df_loc.columns if c not in cols_excluir]
    
    col_widths_map = {
        "DNI": 55, "Locador": 130,
        "Honorario Base": 70, "Días no Prestados": 60,
        "Días Laborados": 60, "Descuento Días": 65, 
        "Otros Pagos": 65, "Pago Bruto": 65, 
        "Retención 4ta (8%)": 72, "Otros Descuentos": 70, 
        "NETO A PAGAR": 70,
    }
    col_w = [col_widths_map.get(c, 60) for c in cols]
    total_w = sum(col_w)
    col_w = [w * W_PAGE / total_w for w in col_w]

    hdr_style = ParagraphStyle('HDR', fontName="Helvetica-Bold", fontSize=7,
                               textColor=colors.white, alignment=1, leading=8)
    nom_style  = ParagraphStyle('NOM', fontName="Helvetica", fontSize=7,
                                textColor=colors.black, alignment=0, leading=8, wordWrap='LTR')
    nom_idx = cols.index("Locador") if "Locador" in cols else -1
    data_rows = [[Paragraph(c, hdr_style) for c in cols]]

    # Variables de fecha para el cálculo de días reales
    mes_num  = int(periodo_key[:2])
    anio_num = int(periodo_key[3:])
    dias_mes = calendar.monthrange(anio_num, mes_num)[1]

    for _, row in df_loc[cols].iterrows():
        fila = []
        dni_l = str(row.get("DNI", ""))
        
        # Lógica de días laborados reales igual que tesorería
        d_vinc = dias_mes
        try:
            db_v = SessionLocal()
            l_obj = db_v.query(Trabajador).filter_by(num_doc=dni_l).first()
            if l_obj and l_obj.fecha_ingreso and l_obj.fecha_ingreso.year == anio_num and l_obj.fecha_ingreso.month == mes_num:
                d_vinc = dias_mes - l_obj.fecha_ingreso.day + 1
            
            v_obj = db_v.query(VariablesMes).filter_by(trabajador_id=l_obj.id, periodo_key=periodo_key).first()
            db_v.close()
            
            susp_t21 = json.loads(v_obj.suspensiones_json or '{}') if v_obj else {}
            cods_desc = ["01", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31", "32", "33", "34", "35"]
            t_susp = sum(int(v or 0) for k, v in susp_t21.items() if k in cods_desc)
        except:
            t_susp = int(row.get("Días no Prestados", 0) or 0)

        d_lab_real = max(0, d_vinc - t_susp)

        for i, c_name in enumerate(cols):
            val = row[c_name]
            if c_name == "Locador":
                fila.append(Paragraph(str(val) if str(val) != "nan" else "", nom_style))
            elif c_name == "Días Laborados":
                fila.append(str(d_lab_real))
            elif isinstance(val, float):
                fila.append(f"{val:,.2f}")
            else:
                fila.append(str(val) if str(val) != "nan" else "")
        data_rows.append(fila)

    # Fila de totales
    tot_style = ParagraphStyle('TC', fontName="Helvetica-Bold", fontSize=7, textColor=colors.white, alignment=1)
    totales_row = []
    cols_texto_loc = {"DNI", "Locador"}
    for c in cols:
        if c == "Locador":
            totales_row.append(Paragraph("<b>TOTALES</b>", tot_style))
        elif c in cols_texto_loc:
            totales_row.append("")
        else:
            try:
                v = df_loc[c].sum()
                totales_row.append(Paragraph(f"<b>{v:,.2f}</b>", tot_style))
            except Exception:
                totales_row.append("")
    data_rows.append(totales_row)

    t = Table(data_rows, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  C_STEEL),
        ('TEXTCOLOR',     (0,0),  (-1,0),  colors.white),
        ('ALIGN',         (0,0),  (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (1,1),  (1,-1),  'LEFT'),
        ('FONTNAME',      (0,1),  (-1,-2), 'Helvetica'),
        ('FONTSIZE',      (0,1),  (-1,-2), 7),
        ('TOPPADDING',    (0,0),  (-1,-1), 4),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 4),
        ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, C_LIGHT]),
        ('BACKGROUND',    (0,-1), (-1,-1), C_NAVY),
        ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,-1), (-1,-1), 7),
        ('GRID',          (0,0),  (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ('LINEABOVE',     (0,-1), (-1,-1), 0.8, C_NAVY),
        ('LINEBELOW',     (0,0),  (-1,0),  0.8, C_GOLD),
    ]))
    elements.append(t)

    # Resumen de retenciones 4ta
    elements.append(Spacer(1, 20))
    st_res = ParagraphStyle('R', fontName="Helvetica-Bold", fontSize=9, textColor=C_NAVY, spaceAfter=6)
    elements.append(Paragraph("RESUMEN DE RETENCIONES — 4ta CATEGORÍA (Ley SUNAT)", st_res))
    try:
        total_bruto = df_loc["Pago Bruto"].sum()
        total_ret   = df_loc["Retención 4ta (8%)"].sum()
        total_neto  = df_loc["NETO A PAGAR"].sum()
        n_loc       = len(df_loc)
        n_con_ret   = len(df_loc[df_loc["Retención 4ta (8%)"] > 0])
        resumen = [
            ["CONCEPTO", "DETALLE", "MONTO (S/)"],
            ["Pago Bruto Total Locadores", f"{n_loc} persona(s)", f"{total_bruto:,.2f}"],
            ["Retención 4ta Categoría (8%)", f"{n_con_ret} locador(es) afecto(s)", f"{total_ret:,.2f}"],
            ["TOTAL NETO A PAGAR", "", f"S/ {total_neto:,.2f}"],
        ]
        t_res = Table(resumen, colWidths=[260, 200, 140])
        estilos_res = [
            ('BACKGROUND', (0,0),  (-1,0),  colors.HexColor("#34495E")),
            ('TEXTCOLOR',  (0,0),  (-1,0),  colors.whitesmoke),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#E5E7E9")),
            ('FONTNAME',   (0,0),  (-1,0),  'Helvetica-Bold'),
            ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0),  (-1,-1), 8),
            ('ALIGN',      (2,0),  (2,-1),  'RIGHT'),
            ('ALIGN',      (0,0),  (1,-1),  'LEFT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('GRID',       (0,0),  (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
        ]
        t_res.setStyle(TableStyle(estilos_res))
        elements.append(t_res)
    except Exception:
        pass

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_pdf_combinado(df_planilla, df_loc, empresa_nombre, periodo_key, empresa_ruc="", empresa_regimen=""):
    """Genera PDF consolidado: Sábana planilla + Valorización locadores + Resumen costo laboral total."""
    periodo_texto = _periodo_legible_calc(periodo_key)
    W_PAGE = 1008 - 24
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=15, bottomMargin=15
    )
    elements = []
    C_NAVY  = colors.HexColor("#0F2744")
    C_STEEL = colors.HexColor("#1E4D8C")
    C_GOLD  = colors.HexColor("#C9A84C")
    C_LIGHT = colors.HexColor("#F0F4F9")
    C_GRAY  = colors.HexColor("#64748B")

    st_title = ParagraphStyle('T',  fontName="Helvetica-Bold", fontSize=13, textColor=C_NAVY, spaceAfter=4)
    st_head  = ParagraphStyle('H',  fontName="Helvetica-Bold", fontSize=10, textColor=C_STEEL, spaceAfter=6)
    st_sec   = ParagraphStyle('SE', fontName="Helvetica-Bold", fontSize=9,  textColor=C_NAVY, spaceAfter=4, spaceBefore=10,
                              borderPad=3, backColor=colors.HexColor("#E8F0FE"))
    st_sub   = ParagraphStyle('S',  fontName="Helvetica",      fontSize=8,  textColor=C_GRAY, spaceAfter=1)

    fecha_calc = datetime.now().strftime("%d/%m/%Y %H:%M")
    ruc_line = f"  |  RUC: {empresa_ruc}" if empresa_ruc else ""
    elements.append(Paragraph(empresa_nombre + ruc_line, st_title))
    elements.append(Paragraph(
        f"REPORTE CONSOLIDADO DE COSTO LABORAL  ·  PERIODO: {periodo_texto}", st_head
    ))
    elements.append(Paragraph(
        f"Planilla de Remuneraciones (5ta Cat.) + Locadores de Servicio (4ta Cat.)  |  Generado: {fecha_calc}", st_sub
    ))
    elements.append(Spacer(1, 8))

    hdr_s = ParagraphStyle('HS', fontName="Helvetica-Bold", fontSize=6, textColor=colors.white, alignment=1, leading=7)
    nom_s = ParagraphStyle('NS', fontName="Helvetica", fontSize=6, textColor=colors.black, alignment=0, leading=7, wordWrap='LTR')

    # ── SECCIÓN 1: PLANILLA ───────────────────────────────────────────────────
    elements.append(Paragraph("▶  1. PLANILLA DE REMUNERACIONES (5ta Categoría)", st_sec))
    _BANCO_COLS_COMB = _COLS_OCULTAS_SABANA | {"Banco", "N° Cuenta", "CCI", "Observaciones"}
    cols_plan = [c for c in df_planilla.columns if c not in _BANCO_COLS_COMB]
    df_p = df_planilla[cols_plan]
    col_widths_plan = {
        "N°": 18, "DNI": 52, "Apellidos y Nombres": 105,
        "Sist. Pensión": 55, "Seg. Social": 48,
        "Sueldo Base": 50, "Asig. Fam.": 36, "Otros Ingresos": 50,
        "TOTAL BRUTO": 50, "ONP (13%)": 42, "AFP Aporte": 42,
        "AFP Seguro": 42, "AFP Comis.": 42, "Ret. 5ta Cat.": 44,
        "Dsctos/Faltas": 44, "NETO A PAGAR": 52, "Aporte Seg. Social": 55,
    }
    col_w_p = [col_widths_plan.get(c, 46) for c in cols_plan]
    total_w_p = sum(col_w_p)
    col_w_p = [w * W_PAGE / total_w_p for w in col_w_p]
    nom_idx_p = cols_plan.index("Apellidos y Nombres") if "Apellidos y Nombres" in cols_plan else -1
    rows_p = [[Paragraph(c, hdr_s) for c in cols_plan]]
    for _, row in df_p.iterrows():
        fila = []
        for i, val in enumerate(row):
            if i == nom_idx_p:
                fila.append(Paragraph(str(val) if str(val) != "nan" else "", nom_s))
            elif isinstance(val, float):
                fila.append(f"{val:,.2f}")
            else:
                fila.append(str(val) if str(val) != "nan" else "")
        rows_p.append(fila)
    t_p = Table(rows_p, colWidths=col_w_p, repeatRows=1)
    t_p.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  C_STEEL),
        ('ALIGN',         (0,0),  (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (2,1),  (2,-1),  'LEFT'),
        ('FONTNAME',      (0,1),  (-1,-2), 'Helvetica'),
        ('FONTSIZE',      (0,1),  (-1,-2), 6),
        ('TOPPADDING',    (0,0),  (-1,-1), 3),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 3),
        ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, C_LIGHT]),
        ('BACKGROUND',    (0,-1), (-1,-1), colors.HexColor("#CBD5E1")),
        ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,-1), (-1,-1), 6),
        ('GRID',          (0,0),  (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ('LINEBELOW',     (0,0),  (-1,0),  0.8, C_GOLD),
    ]))
    elements.append(t_p)

    # ── SECCIÓN 2: LOCADORES ──────────────────────────────────────────────────
    if df_loc is not None and not df_loc.empty:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("▶  2. VALORIZACIÓN DE LOCADORES DE SERVICIO (4ta Categoría)", st_sec))
        # Excluir columnas bancarias y observaciones del reporte combinado
        _BANCO_COMB = {"Banco", "N° Cuenta", "CCI", "Observaciones"}
        cols_loc = [c for c in df_loc.columns if c not in _BANCO_COMB]
        col_widths_loc = {
            "DNI": 52, "Locador": 120,
            "Honorario Base": 65, "Días no Prestados": 55,
            "Descuento Días": 60, "Otros Pagos": 60,
            "Pago Bruto": 60, "Retención 4ta (8%)": 65,
            "Otros Descuentos": 65, "NETO A PAGAR": 65,
        }
        col_w_l = [col_widths_loc.get(c, 55) for c in cols_loc]
        total_w_l = sum(col_w_l)
        col_w_l = [w * W_PAGE / total_w_l for w in col_w_l]
        nom_idx_l = cols_loc.index("Locador") if "Locador" in cols_loc else -1
        rows_l = [[Paragraph(c, hdr_s) for c in cols_loc]]
        for _, row in df_loc[cols_loc].iterrows():
            fila = []
            for i, val in enumerate(row):
                if i == nom_idx_l:
                    fila.append(Paragraph(str(val) if str(val) != "nan" else "", nom_s))
                elif isinstance(val, float):
                    fila.append(f"{val:,.2f}")
                else:
                    fila.append(str(val) if str(val) != "nan" else "")
            rows_l.append(fila)
        t_l = Table(rows_l, colWidths=col_w_l, repeatRows=1)
        t_l.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),  (-1,0),  C_STEEL),
            ('ALIGN',         (0,0),  (-1,-1), 'CENTER'),
            ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
            ('ALIGN',         (1,1),  (1,-1),  'LEFT'),
            ('FONTNAME',      (0,1),  (-1,-1), 'Helvetica'),
            ('FONTSIZE',      (0,1),  (-1,-1), 6),
            ('TOPPADDING',    (0,0),  (-1,-1), 3),
            ('BOTTOMPADDING', (0,0),  (-1,-1), 3),
            ('ROWBACKGROUNDS',(0,1),  (-1,-1), [colors.white, C_LIGHT]),
            ('GRID',          (0,0),  (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
            ('LINEBELOW',     (0,0),  (-1,0),  0.8, C_GOLD),
        ]))
        elements.append(t_l)

    # ── SECCIÓN 3: RESUMEN COSTO LABORAL TOTAL ────────────────────────────────
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("▶  3. RESUMEN CONSOLIDADO DE COSTO LABORAL", st_sec))
    try:
        df_plan_data = df_planilla[df_planilla.get('Apellidos y Nombres', pd.Series(dtype=str)) != 'TOTALES'] \
            if 'Apellidos y Nombres' in df_planilla.columns else df_planilla.iloc[:-1]
        bruto_plan   = df_plan_data['TOTAL BRUTO'].sum()       if 'TOTAL BRUTO'       in df_plan_data.columns else 0.0
        neto_plan    = df_plan_data['NETO A PAGAR'].sum()      if 'NETO A PAGAR'      in df_plan_data.columns else 0.0
        essalud_plan = df_plan_data['Aporte Seg. Social'].sum() if 'Aporte Seg. Social' in df_plan_data.columns else 0.0
        bruto_loc = df_loc["Pago Bruto"].sum()       if df_loc is not None and not df_loc.empty and "Pago Bruto"       in df_loc.columns else 0.0
        neto_loc  = df_loc["NETO A PAGAR"].sum()     if df_loc is not None and not df_loc.empty and "NETO A PAGAR"     in df_loc.columns else 0.0
        ret_4ta   = df_loc["Retención 4ta (8%)"].sum() if df_loc is not None and not df_loc.empty and "Retención 4ta (8%)" in df_loc.columns else 0.0
        costo_total = bruto_plan + essalud_plan + bruto_loc
        neto_total  = neto_plan + neto_loc

        res_data = [
            ["CONCEPTO", "PLANILLA (5ta Cat.)", "LOCADORES (4ta Cat.)", "TOTAL CONSOLIDADO"],
            ["Masa Salarial / Honorarios Brutos",  f"S/ {bruto_plan:,.2f}",             f"S/ {bruto_loc:,.2f}",  f"S/ {(bruto_plan + bruto_loc):,.2f}"],
            ["Aporte EsSalud (cargo empleador)",    f"S/ {essalud_plan:,.2f}",           "—",                     f"S/ {essalud_plan:,.2f}"],
            ["Retenciones PDT (ONP/AFP/5ta/4ta)",  "Ver sábana planilla",               f"S/ {ret_4ta:,.2f}",    "—"],
            ["Total Neto a Pagar al Personal",      f"S/ {neto_plan:,.2f}",             f"S/ {neto_loc:,.2f}",   f"S/ {neto_total:,.2f}"],
            ["COSTO LABORAL TOTAL (empresa)",       f"S/ {bruto_plan + essalud_plan:,.2f}", f"S/ {bruto_loc:,.2f}", f"S/ {costo_total:,.2f}"],
        ]
        t_res = Table(res_data, colWidths=[200, 145, 145, 165])
        t_res.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),  (-1,0),  colors.HexColor("#34495E")),
            ('TEXTCOLOR',     (0,0),  (-1,0),  colors.whitesmoke),
            ('BACKGROUND',    (0,-1), (-1,-1), C_NAVY),
            ('TEXTCOLOR',     (0,-1), (-1,-1), colors.white),
            ('FONTNAME',      (0,0),  (-1,0),  'Helvetica-Bold'),
            ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0),  (-1,-1), 8),
            ('ALIGN',         (1,0),  (-1,-1), 'RIGHT'),
            ('ALIGN',         (0,0),  (0,-1),  'LEFT'),
            ('BOTTOMPADDING', (0,0),  (-1,-1), 6),
            ('TOPPADDING',    (0,0),  (-1,-1), 5),
            ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, C_LIGHT]),
            ('GRID',          (0,0),  (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
        ]))
        elements.append(t_res)
    except Exception:
        pass

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_pdf_tesoreria(df_planilla, df_loc, empresa_nombre, periodo_key, auditoria_data=None, empresa_ruc=""):
    """Genera PDF de Tesorería (Landscape Legal): planilla con columnas de ingresos dinámicas + locadores con datos bancarios."""
    import calendar as _cal_teso
    periodo_texto = _periodo_legible_calc(periodo_key)
    W_PAGE = 1008 - 24

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=95, bottomMargin=15
    )
    elements = []

    C_NAVY  = colors.HexColor("#0F2744")
    C_STEEL = colors.HexColor("#1E4D8C")
    C_GOLD  = colors.HexColor("#C9A84C")
    C_LIGHT = colors.HexColor("#F0F4F9")

    st_sub   = ParagraphStyle('TS',  fontName="Helvetica",      fontSize=9,  textColor=colors.HexColor("#64748B"), spaceAfter=2)
    st_sec   = ParagraphStyle('TSC', fontName="Helvetica-Bold", fontSize=9,  textColor=C_STEEL, spaceBefore=8, spaceAfter=4)
    hdr_s    = ParagraphStyle('TH',  fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.white,  alignment=TA_CENTER, wordWrap='CJK')
    nom_s    = ParagraphStyle('TN',  fontName="Helvetica",      fontSize=6.5, textColor=colors.black,  wordWrap='CJK')
    tot_s    = ParagraphStyle('TTOT',fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.white,  alignment=TA_CENTER)
    st_obs   = ParagraphStyle('OBS', fontName="Helvetica",      fontSize=7.5, textColor=C_NAVY, leading=11, leftIndent=10, spaceAfter=2)

    _fecha_emision = datetime.now().strftime('%d/%m/%Y %H:%M')

    def draw_header(canvas_obj, doc_obj):
        """Membrete fijo en canvas — se repite en cada página."""
        canvas_obj.saveState()
        page_w, page_h = landscape(legal)
        x0, x1 = 12, page_w - 12
        offset_ruc = 14 if empresa_ruc else 0

        # Empresa nombre
        canvas_obj.setFont("Helvetica-Bold", 13)
        canvas_obj.setFillColor(C_NAVY)
        canvas_obj.drawString(x0, page_h - 22, empresa_nombre.upper())

        # Página (derecha)
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.setFillColor(colors.HexColor("#64748B"))
        canvas_obj.drawRightString(x1, page_h - 22, f"Pág. {doc_obj.page}")

        # RUC
        if empresa_ruc:
            canvas_obj.setFont("Helvetica", 9)
            canvas_obj.setFillColor(colors.HexColor("#64748B"))
            canvas_obj.drawString(x0, page_h - 36, f"RUC: {empresa_ruc}")

        # Título del reporte
        canvas_obj.setFont("Helvetica-Bold", 9)
        canvas_obj.setFillColor(C_STEEL)
        canvas_obj.drawString(x0, page_h - 36 - offset_ruc,
                              f"REPORTE DE TESORERÍA — PAGOS DE NÓMINA  |  Periodo: {periodo_texto}")

        # Fecha de emisión
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.setFillColor(colors.HexColor("#64748B"))
        canvas_obj.drawString(x0, page_h - 50 - offset_ruc, f"Emitido el {_fecha_emision}")

        # Líneas separadoras corporativas
        canvas_obj.setStrokeColor(C_NAVY)
        canvas_obj.setLineWidth(1.2)
        canvas_obj.line(x0, page_h - 63 - offset_ruc, x1, page_h - 63 - offset_ruc)
        canvas_obj.setStrokeColor(C_GOLD)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(x0, page_h - 67 - offset_ruc, x1, page_h - 67 - offset_ruc)

        canvas_obj.restoreState()

    # ── TABLA 1: PLANILLA ────────────────────────────────────────────────────────
    elements.append(Paragraph("▶  1. PLANILLA DE REMUNERACIONES (5ta Categoría)", st_sec))

    if df_planilla is not None and not df_planilla.empty:
        df_plan_data = df_planilla[df_planilla.get('Apellidos y Nombres', pd.Series(dtype=str)) != 'TOTALES'] \
            if 'Apellidos y Nombres' in df_planilla.columns else df_planilla.iloc[:-1]

        # Detectar CUALQUIER ingreso que perciba el trabajador de planilla (dinámico total)
        dynamic_income = []
        # 1. Ingresos fijos del DF
        for src, dsp in [("Sueldo Base", "Sueldo Básico"), ("Asig. Fam.", "Asig. Fam.")]:
            if src in df_plan_data.columns and pd.to_numeric(df_plan_data[src], errors='coerce').sum() > 0:
                dynamic_income.append((src, dsp))
        
        # 2. Todos los conceptos dinámicos desde auditoria_data que tengan monto > 0
        all_audit_incomes = set()
        if auditoria_data:
            for dni, info in auditoria_data.items():
                for concepto_nombre, monto in info.get('ingresos', {}).items():
                    if float(monto or 0) > 0 and not concepto_nombre.startswith("Sueldo Base"):
                        all_audit_incomes.add(concepto_nombre)
        
        for inc_name in sorted(list(all_audit_incomes)):
            # Evitar duplicar Asig. Fam si ya se incluyó arriba
            if inc_name != "Asignación Familiar":
                dynamic_income.append((inc_name, inc_name))

        has_5ta   = "Ret. 5ta Cat."  in df_plan_data.columns and df_plan_data["Ret. 5ta Cat."].sum() > 0
        has_dscto = "Dsctos/Faltas"  in df_plan_data.columns and df_plan_data["Dsctos/Faltas"].sum() > 0
        
        # Verificar si hay datos bancarios en planilla
        has_bank_p = df_plan_data["N° Cuenta"].astype(str).str.strip().replace(['nan', 'None', ''], pd.NA).dropna().any() or \
                     df_plan_data["CCI"].astype(str).str.strip().replace(['nan', 'None', ''], pd.NA).dropna().any()

        headers_p = ["N°", "DNI", "Nombres y Apellidos"]
        for _, dsp in dynamic_income:
            headers_p.append(dsp)
        headers_p += ["TOTAL BRUTO", "Ret. Pensiones"]
        if has_5ta:
            headers_p.append("Ret. 5ta Cat.")
        if has_dscto:
            headers_p.append("Otros Dsctos")
        headers_p += ["NETO A PAGAR", "Banco"]
        if has_bank_p:
            headers_p += ["N° Cuenta", "CCI"]

        col_w_map = {
            "N°": 20, "DNI": 48, "Nombres y Apellidos": 110,
            "Sueldo Básico": 58, "Asig. Fam.": 50,
            "Horas Extra 25%": 48, "Horas Extra 35%": 48,
            "Gratificación": 58, "Bono Ext. 9%": 50,
            "TOTAL BRUTO": 62, "Ret. Pensiones": 62,
            "Ret. 5ta Cat.": 52, "Otros Dsctos": 52,
            "NETO A PAGAR": 62, "Banco": 55, "N° Cuenta": 72, "CCI": 82,
        }
        col_w_p = [col_w_map.get(h, 55) for h in headers_p]
        total_wp = sum(col_w_p)
        col_w_p = [w * W_PAGE / total_wp for w in col_w_p]

        rows_p = [[Paragraph(h, hdr_s) for h in headers_p]]
        tot_p = {h: 0.0 for h in headers_p}
        for h in ["N°", "DNI", "Nombres y Apellidos", "Banco", "N° Cuenta", "CCI"]:
            tot_p[h] = None  # texto → no sumar

        for i, (_, row) in enumerate(df_plan_data.iterrows()):
            ret_pens = sum(float(row.get(c, 0.0) or 0.0) for c in ["ONP (13%)", "AFP Aporte", "AFP Seguro", "AFP Comis."])
            fila = [str(i + 1), str(row.get("DNI", "")), Paragraph(str(row.get("Apellidos y Nombres", "")), nom_s)]
            for src, dsp in dynamic_income:
                val = 0.0
                if src in row.index:
                    val = float(row.get(src, 0.0) or 0.0)
                elif auditoria_data:
                    val = float((auditoria_data.get(str(row.get("DNI", "")), {}).get('ingresos', {}) or {}).get(src, 0.0) or 0.0)
                fila.append(f"{val:,.2f}")
                tot_p[dsp] = tot_p.get(dsp, 0.0) + val
            bruto = float(row.get("TOTAL BRUTO", 0.0) or 0.0)
            neto  = float(row.get("NETO A PAGAR", 0.0) or 0.0)
            fila += [f"{bruto:,.2f}", f"{ret_pens:,.2f}"]
            tot_p["TOTAL BRUTO"] = tot_p.get("TOTAL BRUTO", 0.0) + bruto
            tot_p["Ret. Pensiones"] = tot_p.get("Ret. Pensiones", 0.0) + ret_pens
            if has_5ta:
                v5 = float(row.get("Ret. 5ta Cat.", 0.0) or 0.0)
                fila.append(f"{v5:,.2f}")
                tot_p["Ret. 5ta Cat."] = tot_p.get("Ret. 5ta Cat.", 0.0) + v5
            if has_dscto:
                vd = float(row.get("Dsctos/Faltas", 0.0) or 0.0)
                fila.append(f"{vd:,.2f}")
                tot_p["Otros Dsctos"] = tot_p.get("Otros Dsctos", 0.0) + vd
            fila += [f"{neto:,.2f}", str(row.get("Banco", "") or "")]
            if has_bank_p:
                fila += [str(row.get("N° Cuenta", "") or ""), str(row.get("CCI", "") or "")]
            tot_p["NETO A PAGAR"] = tot_p.get("NETO A PAGAR", 0.0) + neto
            rows_p.append(fila)

        # Fila de totales: N° y DNI vacíos, "TOTALES" en columna de nombres (índice 2)
        tot_fila = ["", "", Paragraph("TOTALES", tot_s)]
        for _, dsp in dynamic_income:
            tot_fila.append(f"{tot_p.get(dsp, 0.0):,.2f}")
        tot_fila += [f"{tot_p.get('TOTAL BRUTO', 0.0):,.2f}", f"{tot_p.get('Ret. Pensiones', 0.0):,.2f}"]
        if has_5ta:
            tot_fila.append(f"{tot_p.get('Ret. 5ta Cat.', 0.0):,.2f}")
        if has_dscto:
            tot_fila.append(f"{tot_p.get('Otros Dsctos', 0.0):,.2f}")
        tot_fila += [f"{tot_p.get('NETO A PAGAR', 0.0):,.2f}", ""]
        if has_bank_p:
            tot_fila += ["", ""]
        rows_p.append(tot_fila)

        t1 = Table(rows_p, colWidths=col_w_p, repeatRows=1)
        t1.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0),  (-1, 0),  C_NAVY),
            ('BACKGROUND',    (0, -1), (-1, -1), C_STEEL),
            ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
            ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
            ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
            ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
            ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0),  (-1, -1), 6.5),
            ('ALIGN',         (0, 0),  (-1, -1), 'CENTER'),
            ('ALIGN',         (2, 1),  (2, -2),  'LEFT'),
            ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0),  (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0),  (-1, -1), 3),
            ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, C_LIGHT]),
            ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ('LINEBELOW',     (0, 0),  (-1, 0),  0.8, C_GOLD),
        ]))
        elements.append(t1)
    else:
        elements.append(Paragraph("(Sin datos de planilla para este periodo)", st_sub))

    # ── TABLA 2: LOCADORES ────────────────────────────────────────────────────────
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("▶  2. LOCADORES DE SERVICIO (4ta Categoría)", st_sec))

    if df_loc is not None and not df_loc.empty:
        mes_num  = int(periodo_key[:2])
        anio_num = int(periodo_key[3:])
        dias_mes = _cal_teso.monthrange(anio_num, mes_num)[1]
        has_dias_lab = "Días Laborados" in df_loc.columns
        
        # Verificar si hay datos bancarios en locadores
        has_bank_l = df_loc["N° Cuenta"].astype(str).str.strip().replace(['nan', 'None', ''], pd.NA).dropna().any() or \
                     df_loc["CCI"].astype(str).str.strip().replace(['nan', 'None', ''], pd.NA).dropna().any()

        headers_l = ["N°", "DNI", "Nombres y Apellidos", "Honorario Base"]
        if has_dias_lab:
            headers_l.append("Días Laborados")
        headers_l += ["Pago Bruto", "Retención 4ta", "Otros Dsctos", "NETO A PAGAR", "Banco"]
        if has_bank_l:
            headers_l += ["N° Cuenta", "CCI"]

        col_w_lmap = {
            "N°": 20, "DNI": 48, "Nombres y Apellidos": 110,
            "Honorario Base": 60, "Días Laborados": 48,
            "Pago Bruto": 60, "Retención 4ta": 60, "Otros Dsctos": 52, "NETO A PAGAR": 60,
            "Banco": 52, "N° Cuenta": 70, "CCI": 80,
        }
        col_w_l = [col_w_lmap.get(h, 55) for h in headers_l]
        total_wl = sum(col_w_l)
        col_w_l = [w * W_PAGE / total_wl for w in col_w_l]

        loc_col = "Locador" if "Locador" in df_loc.columns else df_loc.columns[2]
        rows_l = [[Paragraph(h, hdr_s) for h in headers_l]]
        tot_hon = tot_bruto = tot_ret = tot_dscto = tot_neto = 0.0

        for i, (_, row) in enumerate(df_loc.iterrows()):
            # Ajuste visual para Tesorería: días efectivamente laborados
            # 1. Días según fecha de ingreso vs calendario real
            dni_l = str(row.get("DNI", ""))
            dias_vinc = dias_mes
            try:
                db_v = SessionLocal()
                l_obj = db_v.query(Trabajador).filter_by(num_doc=dni_l).first()
                if l_obj and l_obj.fecha_ingreso and l_obj.fecha_ingreso.year == anio_num and l_obj.fecha_ingreso.month == mes_num:
                    dias_vinc = dias_mes - l_obj.fecha_ingreso.day + 1
                
                # 2. Descontar suspensiones Tabla 21
                v_obj = db_v.query(VariablesMes).filter_by(trabajador_id=l_obj.id, periodo_key=periodo_key).first()
                db_v.close()
                
                susp_t21 = json.loads(v_obj.suspensiones_json or '{}') if v_obj else {}
                cods_desc = ["01", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31", "32", "33", "34", "35"]
                total_susp = sum(int(v or 0) for k, v in susp_t21.items() if k in cods_desc)
            except:
                total_susp = int(row.get("Días no Prestados", 0) or 0)

            dias_lab = max(0, dias_vinc - total_susp)
            
            hon_b  = float(row.get("Honorario Base", 0.0) or 0.0)
            bruto  = float(row.get("Pago Bruto", 0.0) or 0.0)
            ret    = float(row.get("Retención 4ta (8%)", 0.0) or 0.0)
            dscto  = float(row.get("Otros Descuentos", 0.0) or 0.0)
            neto   = float(row.get("NETO A PAGAR", 0.0) or 0.0)
            fila   = [str(i + 1), str(row.get("DNI", "") or ""),
                      Paragraph(str(row.get(loc_col, "") or ""), nom_s),
                      f"{hon_b:,.2f}"]
            if has_dias_lab:
                fila.append(str(dias_lab))
            fila += [
                f"{bruto:,.2f}", f"{ret:,.2f}", f"{dscto:,.2f}", f"{neto:,.2f}",
                str(row.get("Banco", "") or ""),
            ]
            if has_bank_l:
                fila += [str(row.get("N° Cuenta", "") or ""), str(row.get("CCI", "") or "")]
            rows_l.append(fila)
            tot_hon += hon_b; tot_bruto += bruto; tot_ret += ret; tot_dscto += dscto; tot_neto += neto

        # Fila de totales locadores: "TOTALES" en columna de nombres (índice 2)
        tot_l_fila = ["", "", Paragraph("TOTALES", tot_s), f"{tot_hon:,.2f}"]
        if has_dias_lab:
            tot_l_fila.append("")
        tot_l_fila += [f"{tot_bruto:,.2f}", f"{tot_ret:,.2f}", f"{tot_dscto:,.2f}", f"{tot_neto:,.2f}", ""]
        if has_bank_l:
            tot_l_fila += ["", ""]
        rows_l.append(tot_l_fila)

        t2 = Table(rows_l, colWidths=col_w_l, repeatRows=1)
        t2.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0),  (-1, 0),  C_NAVY),
            ('BACKGROUND',    (0, -1), (-1, -1), C_STEEL),
            ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
            ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
            ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
            ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
            ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0),  (-1, -1), 6.5),
            ('ALIGN',         (0, 0),  (-1, -1), 'CENTER'),
            ('ALIGN',         (2, 1),  (2, -2),  'LEFT'),
            ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0),  (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0),  (-1, -1), 3),
            ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, C_LIGHT]),
            ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ('LINEBELOW',     (0, 0),  (-1, 0),  0.8, C_GOLD),
        ]))
        elements.append(t2)
    else:
        elements.append(Paragraph("(Sin locadores de servicio para este periodo)", st_sub))

    # ── SECCIÓN 3: OBSERVACIONES DEL PERIODO ─────────────────────────────────────
    obs_all = []
    if auditoria_data:
        for _dni_o, _info_o in auditoria_data.items():
            obs_str = _info_o.get('observaciones', '')
            if obs_str:
                nombre_o = _info_o.get('nombres', str(_dni_o))
                obs_all.append(f"[Planilla] {nombre_o} ({_dni_o}): {obs_str}")
    if df_loc is not None and not df_loc.empty and 'Observaciones' in df_loc.columns:
        _loc_col_o = "Locador" if "Locador" in df_loc.columns else df_loc.columns[2]
        for _, row_o in df_loc.iterrows():
            obs_str = str(row_o.get('Observaciones', '') or '')
            if obs_str:
                obs_all.append(f"[Locador] {str(row_o.get(_loc_col_o, ''))} ({str(row_o.get('DNI', ''))}): {obs_str}")

    if obs_all:
        elements.append(Spacer(1, 14))
        elements.append(Paragraph("▶  3. OBSERVACIONES DEL PERIODO", st_sec))
        for obs_line in obs_all:
            elements.append(Paragraph(f"• {obs_line}", st_obs))

    doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header)
    buffer.seek(0)
    return buffer


def generar_pdf_personalizado(df, empresa_nombre, periodo_key, titulo, empresa_ruc=""):
    """
    Genera un PDF corporativo para el Reporte Personalizado con cualquier combinación
    de columnas seleccionadas por el usuario.
    Paleta corporativa C_NAVY / C_STEEL / C_GOLD. Landscape Legal.
    Encabezado corporativo repetido en canvas (topMargin=95).
    Si "Observaciones" está en las columnas, se imprime como sección al final (no en la tabla).
    """
    periodo_texto = _periodo_legible_calc(periodo_key)
    W_PAGE = 1008 - 24  # 984 pt disponibles en landscape legal con márgenes 12

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=95, bottomMargin=15
    )
    elements = []

    C_NAVY  = colors.HexColor("#0F2744")
    C_STEEL = colors.HexColor("#1E4D8C")
    C_GOLD  = colors.HexColor("#C9A84C")
    C_LIGHT = colors.HexColor("#F0F4F9")

    st_sec   = ParagraphStyle('RPSC', fontName="Helvetica-Bold", fontSize=9,  textColor=C_STEEL, spaceBefore=8, spaceAfter=4)
    hdr_s    = ParagraphStyle('RPH',  fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.white,
                               alignment=TA_CENTER, wordWrap='CJK', leading=8)
    nom_s    = ParagraphStyle('RPN',  fontName="Helvetica",      fontSize=6.5, textColor=colors.black,
                               wordWrap='CJK', leading=8)
    tot_s    = ParagraphStyle('RPTOT',fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.white,
                               alignment=TA_CENTER)
    st_obs   = ParagraphStyle('OBS',  fontName="Helvetica",      fontSize=7.5, textColor=C_NAVY,
                               leading=11, leftIndent=10, spaceAfter=2)

    _fecha_emision = datetime.now().strftime('%d/%m/%Y %H:%M')

    def draw_header(canvas_obj, doc_obj):
        """Membrete fijo en canvas — se repite en cada página."""
        canvas_obj.saveState()
        page_w, page_h = landscape(legal)
        x0, x1 = 12, page_w - 12
        offset_ruc = 14 if empresa_ruc else 0

        # Empresa nombre
        canvas_obj.setFont("Helvetica-Bold", 13)
        canvas_obj.setFillColor(C_NAVY)
        canvas_obj.drawString(x0, page_h - 22, empresa_nombre.upper())

        # Página (derecha)
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.setFillColor(colors.HexColor("#64748B"))
        canvas_obj.drawRightString(x1, page_h - 22, f"Pág. {doc_obj.page}")

        # RUC
        if empresa_ruc:
            canvas_obj.setFont("Helvetica", 9)
            canvas_obj.setFillColor(colors.HexColor("#64748B"))
            canvas_obj.drawString(x0, page_h - 36, f"RUC: {empresa_ruc}")

        # Título del reporte
        canvas_obj.setFont("Helvetica-Bold", 9)
        canvas_obj.setFillColor(C_STEEL)
        canvas_obj.drawString(x0, page_h - 36 - offset_ruc,
                              f"{titulo.upper()}  |  Periodo: {periodo_texto}")

        # Fecha de emisión
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.setFillColor(colors.HexColor("#64748B"))
        canvas_obj.drawString(x0, page_h - 50 - offset_ruc, f"Emitido el {_fecha_emision}")

        # Líneas separadoras corporativas
        canvas_obj.setStrokeColor(C_NAVY)
        canvas_obj.setLineWidth(1.2)
        canvas_obj.line(x0, page_h - 63 - offset_ruc, x1, page_h - 63 - offset_ruc)
        canvas_obj.setStrokeColor(C_GOLD)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(x0, page_h - 67 - offset_ruc, x1, page_h - 67 - offset_ruc)

        canvas_obj.restoreState()

    if df is None or df.empty:
        doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header)
        buffer.seek(0)
        return buffer

    cols = list(df.columns)

    # "Observaciones" se excluye de la tabla y se renderiza al final como sección
    tiene_obs = "Observaciones" in cols
    cols_tabla = [c for c in cols if c != "Observaciones"]

    # Anchos base por columna (se escalan proporcionalmente a W_PAGE)
    _DEFAULT_W = {
        "N°": 20, "DNI": 48,
        "Apellidos y Nombres": 120, "Locador": 120, "Nombres y Apellidos": 120,
        "Cargo": 75, "Cargo / Actividad": 80,
        "Sist. Pensión": 55, "Seg. Social": 50,
        "Sueldo Base": 52, "Asig. Fam.": 40, "Otros Ingresos": 52,
        "TOTAL BRUTO": 56, "ONP (13%)": 46, "AFP Aporte": 46,
        "AFP Seguro": 46, "AFP Comis.": 46, "Ret. 5ta Cat.": 48,
        "Dsctos/Faltas": 46, "NETO A PAGAR": 58, "Aporte Seg. Social": 58,
        "Honorario Base": 58, "Días no Prestados": 52, "Descuento Días": 52,
        "Otros Pagos": 52, "Pago Bruto": 56, "Retención 4ta (8%)": 56,
        "Otros Descuentos": 56,
        "Banco": 55, "N° Cuenta": 72, "CCI": 82,
    }
    col_w = [_DEFAULT_W.get(c, 58) for c in cols_tabla]
    total_w = sum(col_w)
    col_w = [w * W_PAGE / total_w for w in col_w]

    # Columnas bancarias — se renderizan con Paragraph para word-wrap
    _BANCO_COLS = {"Banco", "N° Cuenta", "CCI"}

    # Columnas de texto (no se suman en la fila de totales)
    _TEXTO = {
        "N°", "DNI", "Apellidos y Nombres", "Locador", "Nombres y Apellidos",
        "Cargo", "Cargo / Actividad", "Sist. Pensión", "Seg. Social",
        "Banco", "N° Cuenta", "CCI",
    }
    # Índice de la columna principal de nombres (para word-wrap y alineación izquierda)
    nom_idx = next(
        (i for i, c in enumerate(cols_tabla) if c in {"Apellidos y Nombres", "Locador", "Nombres y Apellidos"}),
        -1
    )

    # ── Construir filas de la tabla ──────────────────────────────────────────────
    rows_pdf = [[Paragraph(c, hdr_s) for c in cols_tabla]]

    for _, row in df.iterrows():
        fila = []
        for j, c in enumerate(cols_tabla):
            val = row.get(c, "")
            val_str = str(val) if str(val) not in ("nan", "None", "NaN") else ""
            if j == nom_idx or c in _BANCO_COLS:
                fila.append(Paragraph(val_str, nom_s))
            elif isinstance(val, float) and str(val) not in ("nan", "NaN"):
                fila.append(f"{val:,.2f}")
            else:
                fila.append(val_str)
        rows_pdf.append(fila)

    # ── Fila de totales ──────────────────────────────────────────────────────────
    tot_row = []
    for j, c in enumerate(cols_tabla):
        if c in _TEXTO:
            # "TOTALES" va en la columna de nombres; las demás texto quedan vacías
            tot_row.append(Paragraph("TOTALES", tot_s) if j == nom_idx else "")
        else:
            try:
                total = pd.to_numeric(df[c], errors='coerce').fillna(0).sum()
                tot_row.append(f"{total:,.2f}")
            except Exception:
                tot_row.append("")
    # Si no hay columna de nombres, poner TOTALES en índice 2 (o 0 si hay menos columnas)
    if nom_idx < 0 and len(cols_tabla) > 0:
        tot_row[min(2, len(cols_tabla) - 1)] = Paragraph("TOTALES", tot_s)
    rows_pdf.append(tot_row)

    # ── Estilos de la tabla ──────────────────────────────────────────────────────
    style_cmds = [
        ('BACKGROUND',    (0, 0),  (-1, 0),  C_NAVY),
        ('BACKGROUND',    (0, -1), (-1, -1), C_STEEL),
        ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
        ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
        ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0),  (-1, -1), 6.5),
        ('ALIGN',         (0, 0),  (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0),  (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0),  (-1, -1), 3),
        ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, C_LIGHT]),
        ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ('LINEBELOW',     (0, 0),  (-1, 0),  0.8, C_GOLD),
    ]
    if nom_idx >= 0:
        style_cmds.append(('ALIGN', (nom_idx, 1), (nom_idx, -2), 'LEFT'))
    for j, c in enumerate(cols_tabla):
        if c in _BANCO_COLS:
            style_cmds.append(('ALIGN', (j, 1), (j, -2), 'LEFT'))

    t = Table(rows_pdf, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    # ── SECCIÓN: OBSERVACIONES DEL PERIODO ──────────────────────────────────────
    if tiene_obs and 'Observaciones' in df.columns:
        _nom_col = next((c for c in ["Apellidos y Nombres", "Locador", "Nombres y Apellidos"] if c in df.columns), None)
        _dni_col = "DNI" if "DNI" in df.columns else None
        obs_all = []
        for _, row_o in df.iterrows():
            obs_str = str(row_o.get('Observaciones', '') or '').strip()
            if obs_str:
                nombre_o = str(row_o.get(_nom_col, '')) if _nom_col else ''
                dni_o    = str(row_o.get(_dni_col, '')) if _dni_col else ''
                label    = f"{nombre_o} ({dni_o})" if dni_o else nombre_o
                obs_all.append(f"{label}: {obs_str}")
        if obs_all:
            elements.append(Spacer(1, 14))
            elements.append(Paragraph("▶  OBSERVACIONES DEL PERIODO", st_sec))
            for obs_line in obs_all:
                elements.append(Paragraph(f"• {obs_line}", st_obs))

    doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header)
    buffer.seek(0)
    return buffer


# --- 2. MOTOR DE RENDERIZADO Y CÁLCULO ---

def _render_planilla_tab(empresa_id, empresa_nombre, mes_seleccionado, anio_seleccionado, periodo_key, mes_idx):
    # ─── LEER DATOS DESDE NEON ────────────────────────────────────────────────
    db = SessionLocal()
    try:
        # 1. Parámetros Legales
        p = _cargar_parametros(db, empresa_id, periodo_key)
        if not p:
            st.error(f"🛑 ALTO: No se han configurado los Parámetros Legales para el periodo **{periodo_key}**.")
            st.info("Vaya al módulo 'Parámetros Legales' y configure las tasas para este periodo.")
            return

        # 1b. Jornada diaria de la empresa (default 8 h)
        from infrastructure.database.models import Empresa as EmpresaModel
        empresa_obj = db.query(EmpresaModel).filter_by(id=empresa_id).first()
        horas_jornada = float(getattr(empresa_obj, 'horas_jornada_diaria', None) or 8.0)

        # 2. Trabajadores activos
        df_trab = _cargar_trabajadores_df(db, empresa_id)
        if df_trab.empty:
            st.warning("⚠️ No hay trabajadores activos registrados en el Maestro de Personal.")
            return

        # 3. Conceptos de la empresa
        conceptos_list = db.query(Concepto).filter_by(empresa_id=empresa_id).all()
        conceptos_empresa = _cargar_conceptos_df(db, empresa_id)

        # 4. Variables del periodo
        df_var = _cargar_variables_df(db, empresa_id, periodo_key, conceptos_list)
        if df_var.empty:
            st.warning(f"⚠️ No se han ingresado Asistencias para **{periodo_key}**.")
            st.info("Vaya al módulo 'Ingreso de Asistencias' y guarde las variables del mes.")
            return

        # Merge principal (igual que antes)
        df_planilla = pd.merge(df_trab, df_var, on="Num. Doc.", how="inner")

        # Compatibilidad con emision_boletas.py (lee de session_state)
        st.session_state['trabajadores_mock'] = df_trab
        if 'variables_por_periodo' not in st.session_state:
            st.session_state['variables_por_periodo'] = {}
        st.session_state['variables_por_periodo'][periodo_key] = df_var

    finally:
        db.close()

    # ── VERIFICAR ESTADO DE CIERRE ─────────────────────────────────────────
    es_cerrada = False
    try:
        db_ck = SessionLocal()
        plan_ck = db_ck.query(PlanillaMensual).filter_by(
            empresa_id=empresa_id, periodo_key=periodo_key
        ).first()
        db_ck.close()
        if plan_ck and getattr(plan_ck, 'estado', 'ABIERTA') == 'CERRADA':
            es_cerrada = True
    except Exception:
        pass

    # ── VERIFICAR LOCADORES SIN ASISTENCIA GUARDADA ──────────────────────────
    locadores_pendientes = False
    try:
        db_lv = SessionLocal()
        # Solo locadores que ya deberían haber ingresado según su fecha de ingreso
        _mes_c = int(periodo_key[:2])
        _ani_c = int(periodo_key[3:])
        
        locs_activos = db_lv.query(Trabajador).filter_by(
            empresa_id=empresa_id, situacion="ACTIVO", tipo_contrato="LOCADOR"
        ).all()
        
        locs_que_corresponden = [
            l for l in locs_activos 
            if not (l.fecha_ingreso and (l.fecha_ingreso.year > _ani_c or (l.fecha_ingreso.year == _ani_c and l.fecha_ingreso.month > _mes_c)))
        ]
        
        if locs_que_corresponden:
            ids_locs = [l.id for l in locs_que_corresponden]
            n_vars_loc = db_lv.query(VariablesMes).filter(
                VariablesMes.empresa_id == empresa_id,
                VariablesMes.periodo_key == periodo_key,
                VariablesMes.trabajador_id.in_(ids_locs)
            ).count()
            locadores_pendientes = (n_vars_loc < len(locs_que_corresponden))
        db_lv.close()
    except Exception:
        locadores_pendientes = False

    if es_cerrada:
        st.error(f"🔒 La planilla del periodo **{periodo_key}** ya fue CERRADA y contabilizada. Vaya al final de la página para reabrirla si tiene permisos de Supervisor.")
    elif locadores_pendientes:
        st.warning(
            f"⚠️ **CÁLCULO BLOQUEADO:** Hay locadores de servicio sin asistencia guardada para **{periodo_key}**. "
            f"Vaya al módulo **'Ingreso de Asistencias'** → pestaña **'🧾 2. Valorización de Locadores'** "
            f"y guarde antes de ejecutar el motor de planilla."
        )
    elif st.button(f"🚀 Ejecutar Motor de Planilla - {periodo_key}", type="primary", use_container_width=True):
        st.session_state['ultima_planilla_calculada'] = True
        resultados = []
        auditoria_data = {}

        # Cargar historial de quinta categoría de periodos anteriores del mismo año
        historico_quinta: dict = {}
        try:
            db_hq = SessionLocal()
            for mes_ant in range(1, mes_idx):
                periodo_ant = f"{mes_ant:02d}-{anio_seleccionado}"
                plan_ant = db_hq.query(PlanillaMensual).filter_by(
                    empresa_id=empresa_id, periodo_key=periodo_ant
                ).first()
                if plan_ant:
                    try:
                        aud_ant = json.loads(plan_ant.auditoria_json or '{}')
                        for dni_ant, data_ant in aud_ant.items():
                            q_ant = data_ant.get('quinta', {})
                            b = float(q_ant.get('base_mes', 0.0))
                            r = float(q_ant.get('retencion', 0.0))
                            if b > 0 or r > 0:
                                if dni_ant not in historico_quinta:
                                    historico_quinta[dni_ant] = {'rem_previa': 0.0, 'ret_previa': 0.0}
                                historico_quinta[dni_ant]['rem_previa'] += b
                                historico_quinta[dni_ant]['ret_previa'] += r
                    except Exception:
                        pass
            db_hq.close()
        except Exception:
            pass

        # Precargar cuotas de préstamos pendientes del periodo (una sola consulta)
        cuotas_del_mes: dict = {}
        try:
            db_cuotas = SessionLocal()
            _cuotas = (
                db_cuotas.query(CuotaPrestamo)
                .join(Prestamo)
                .filter(
                    Prestamo.empresa_id == empresa_id,
                    CuotaPrestamo.periodo_key == periodo_key,
                    CuotaPrestamo.estado == 'PENDIENTE',
                )
                .all()
            )
            for _c in _cuotas:
                _dni_c = _c.prestamo.trabajador.num_doc
                cuotas_del_mes.setdefault(_dni_c, []).append({
                    'id':            _c.id,
                    'numero_cuota':  _c.numero_cuota,
                    'numero_cuotas': _c.prestamo.numero_cuotas,
                    'concepto':      _c.prestamo.concepto,
                    'monto':         float(_c.monto),
                })
            db_cuotas.close()
        except Exception:
            cuotas_del_mes = {}

        # Precargar notas de gestión manuales
        notas_gestion_map = {}
        try:
            db_n = SessionLocal()
            _v_notas = db_n.query(VariablesMes).filter_by(empresa_id=empresa_id, periodo_key=periodo_key).all()
            for _vn in _v_notas:
                notas_gestion_map[_vn.trabajador.num_doc] = getattr(_vn, 'notas_gestion', '') or ''
            db_n.close()
        except: pass

        for index, row in df_planilla.iterrows():
            # Definir variables de cálculo para filtros y lógica proporcional
            mes_calc  = int(mes_seleccionado[:2])
            anio_calc = int(anio_seleccionado)

            # Filtro de fecha de ingreso para personal en planilla
            try:
                fi_p = pd.to_datetime(row['Fecha Ingreso'])
                if fi_p.year > anio_calc or (fi_p.year == anio_calc and fi_p.month > mes_calc):
                    continue
            except: pass

            dni_trabajador = row['Num. Doc.']
            nombres = row['Nombres y Apellidos_x']
            sistema = str(row.get('Sistema Pensión', 'NO AFECTO')).upper()
            
            # --- TIEMPOS Y BASES FIJAS (Proporcionalidad Segura) ---
            try:
                fecha_ingreso = pd.to_datetime(row['Fecha Ingreso'])

                dias_del_mes = calendar.monthrange(anio_calc, mes_calc)[1]
                ingreso_este_mes = (fecha_ingreso.year == anio_calc and fecha_ingreso.month == mes_calc)
                dias_computables = dias_del_mes
                if ingreso_este_mes:
                    dias_computables = max(0, dias_del_mes - fecha_ingreso.day + 1)
                elif fecha_ingreso.year > anio_calc or (fecha_ingreso.year == anio_calc and fecha_ingreso.month > mes_calc):
                    dias_computables = 0
            except Exception:
                dias_del_mes = 30
                dias_computables = 30
                ingreso_este_mes = False

            # Trabajador aún no ingresa en este periodo — omitir completamente
            if dias_computables == 0:
                continue

            # Suspensiones desde suspensiones_json; fallback a Días Faltados
            susp_raw = str(row.get('suspensiones_json', '{}') or '{}')
            try:
                susp_dict = json.loads(susp_raw)
            except Exception:
                susp_dict = {}
            total_ausencias   = sum(susp_dict.values()) if susp_dict else float(row.get('Días Faltados', 0))
            dias_laborados    = max(0, int(dias_computables) - int(total_ausencias))
            factor_asistencia = dias_laborados / dias_computables if dias_computables > 0 else 0.0
            horas_ordinarias  = int(dias_laborados * horas_jornada)

            sueldo_base_nominal = float(row['Sueldo Base'])
            valor_dia           = sueldo_base_nominal / 30.0   # Base 30 — Mes Comercial Mixto
            valor_hora          = valor_dia / horas_jornada

            # Mes Comercial Mixto: trabajó todos los días disponibles → sueldo íntegro; parcial → proporcional/30
            # (incluye el caso de ingreso día 1: dias_computables == dias_del_mes → mes completo)
            if total_ausencias == 0 and (not ingreso_este_mes or dias_computables >= dias_del_mes):
                sueldo_computable = sueldo_base_nominal
            else:
                sueldo_computable = max(0.0, valor_dia * dias_laborados)

            # --- OBSERVACIONES DEL PERIODO ---
            obs_trab = []
            if ingreso_este_mes:
                obs_trab.append(f"Ingresó el {fecha_ingreso.strftime('%d/%m/%Y')}")
            
            # Detalle de descuentos por ausencias para Tesorería
            monto_dscto_ausencias = round(sueldo_base_nominal - sueldo_computable, 2)
            if total_ausencias > 0:
                obs_trab.append(f"Días no laborados: {int(total_ausencias)} (Desc: S/ {monto_dscto_ausencias:,.2f})")

            dscto_tardanzas = float(row['Min. Tardanza']) * (valor_hora / 60)
            if dscto_tardanzas > 0:
                obs_trab.append(f"Tardanzas: {int(row['Min. Tardanza'])} min (Desc: S/ {dscto_tardanzas:,.2f})")
            
            # Integrar Nota de Gestión Manual
            nota_manual = notas_gestion_map.get(str(dni_trabajador), "")
            if nota_manual:
                obs_trab.append(f"NOTA: {nota_manual}")

            # Asig. familiar: se paga solo si hay sueldo computable > 0 y al menos 1 día remunerado.
            # Códigos remunerados (no descuentan asig.fam): 20=Desc.Médico, 23=Vacaciones, 25=Lic.c/Goce
            _COD_REM = {"20", "23", "25"}
            dias_remunerados = dias_laborados + sum(int(susp_dict.get(c, 0)) for c in _COD_REM)
            tiene_asig_fam = (row.get('Asig. Fam.', "No") == "Sí"
                              and sueldo_computable > 0
                              and dias_remunerados > 0)
            monto_asig_fam = (p['rmv'] * 0.10) if tiene_asig_fam else 0.0

            pago_he_25 = float(row.get('Hrs Extras 25%', 0.0)) * (valor_hora * 1.25)
            pago_he_35 = float(row.get('Hrs Extras 35%', 0.0)) * (valor_hora * 1.35)

            ingresos_totales    = sueldo_computable + monto_asig_fam + pago_he_25 + pago_he_35
            # Solo tardanzas como descuento manual; las faltas ya reducen sueldo_computable
            descuentos_manuales = dscto_tardanzas
            desglose_descuentos = {}
            if dscto_tardanzas > 0:
                desglose_descuentos["Tardanzas"] = round(dscto_tardanzas, 2)

            # ── CUOTAS DE PRÉSTAMOS/DESCUENTOS PROGRAMADOS ──────────────────
            # IMPORTANTE: Estos montos NO afectan bases de AFP, 5ta o EsSalud. 
            # Solo se restan al final para llegar al NETO A PAGAR.
            for _cuota in cuotas_del_mes.get(str(dni_trabajador), []):
                _monto_c = _cuota['monto']
                descuentos_manuales += _monto_c
                _concepto_c = _cuota['concepto']
                desglose_descuentos[_concepto_c] = desglose_descuentos.get(_concepto_c, 0.0) + _monto_c
                obs_trab.append(
                    f"{_concepto_c}: Cuota {_cuota['numero_cuota']}/{_cuota['numero_cuotas']}"
                    f" (S/ {_monto_c:,.2f})"
                )

            # ── OBSERVACIONES ADICIONALES ────────────────────────────────────
            # Verificación bancaria
            if not str(row.get('Banco', '') or '').strip() or not str(row.get('Cuenta Bancaria', '') or '').strip():
                obs_trab.append("⚠️ Sin cuenta bancaria (Pago manual)")
            # Descanso médico (código 20) o accidente de trabajo (código 16)
            for _cod, _desc in [("20", "Descanso médico"), ("16", "Accidente de trabajo")]:
                if int(susp_dict.get(_cod, 0)) > 0:
                    obs_trab.append(f"{_desc}: {int(susp_dict[_cod])} día(s)")

            # Las bases imponibles se calculan ANTES de aplicar descuentos de préstamos
            base_afp_onp        = ingresos_totales
            base_essalud        = ingresos_totales
            base_quinta_mes     = ingresos_totales

            desglose_ingresos = {
                f"Sueldo Base ({int(dias_laborados)} días)": round(sueldo_computable, 2),
                "Asignación Familiar": round(monto_asig_fam, 2),
            }
            if pago_he_25 > 0:
                desglose_ingresos["Horas Extras 25%"] = round(pago_he_25, 2)
            if pago_he_35 > 0:
                desglose_ingresos["Horas Extras 35%"] = round(pago_he_35, 2)

            # --- CONCEPTOS DINÁMICOS Y GRATIFICACIONES ---
            monto_grati = float(row.get('GRATIFICACION (JUL/DIC)', 0.0))
            if monto_grati > 0:
                monto_bono_9 = monto_grati * 0.09
                desglose_ingresos['Gratificación'] = round(monto_grati, 2)
                desglose_ingresos['Bono Ext. 9%'] = round(monto_bono_9, 2)
                ingresos_totales += (monto_grati + monto_bono_9)
                base_quinta_mes += (monto_grati + monto_bono_9)

            conceptos_omitidos = ["SUELDO BASICO", "ASIGNACION FAMILIAR", "GRATIFICACION (JUL/DIC)", "BONIFICACION EXTRAORDINARIA LEY 29351 (9%)"]
            otros_ingresos = 0.0
            conceptos_recuperados_5ta = 0.0
            for _, concepto in conceptos_empresa.iterrows():
                nombre_c = concepto['Nombre del Concepto']
                if nombre_c in conceptos_omitidos: continue
                if nombre_c in row and float(row[nombre_c]) > 0:
                    monto_ingresado_nominal = float(row[nombre_c])
                    if concepto.get('Prorrateable', False):
                        monto_concepto = monto_ingresado_nominal * factor_asistencia
                        # Si es un ingreso afecto a 5ta, guardamos el diferencial que se perdió por faltar
                        if concepto['Tipo'] == "INGRESO" and concepto.get('Afecto 5ta Cat.', False):
                            conceptos_recuperados_5ta += (monto_ingresado_nominal - monto_concepto)
                    else:
                        monto_concepto = monto_ingresado_nominal

                    if concepto['Tipo'] == "INGRESO":
                        desglose_ingresos[nombre_c] = round(monto_concepto, 2)
                        otros_ingresos += monto_concepto
                        ingresos_totales += monto_concepto
                        if concepto['Afecto AFP/ONP']: base_afp_onp += monto_concepto
                        if concepto['Afecto EsSalud']: base_essalud += monto_concepto
                        if concepto['Afecto 5ta Cat.']: base_quinta_mes += monto_concepto
                    elif concepto['Tipo'] == "DESCUENTO":
                        desglose_descuentos[nombre_c] = round(monto_concepto, 2)
                        descuentos_manuales += monto_concepto
                        if concepto['Afecto AFP/ONP']: base_afp_onp -= monto_concepto
                        if concepto['Afecto EsSalud']: base_essalud -= monto_concepto
                        if concepto['Afecto 5ta Cat.']: base_quinta_mes -= monto_concepto

            base_afp_onp = max(0.0, base_afp_onp)
            base_essalud = max(0.0, base_essalud)
            base_quinta_mes = max(0.0, base_quinta_mes)
            
            # --- CÁLCULO DE PENSIONES ---
            aporte_afp = 0.0
            prima_afp = 0.0
            comis_afp = 0.0
            dscto_onp = 0.0
            
            if sistema == "ONP":
                dscto_onp = base_afp_onp * (p['tasa_onp'] / 100)
                if dscto_onp > 0: desglose_descuentos['Aporte ONP'] = round(dscto_onp, 2)
            elif sistema != "NO AFECTO":
                prefijo = ""
                if "HABITAT" in sistema: prefijo = "afp_habitat_"
                elif "INTEGRA" in sistema: prefijo = "afp_integra_"
                elif "PRIMA" in sistema: prefijo = "afp_prima_"
                elif "PROFUTURO" in sistema: prefijo = "afp_profuturo_"

                if prefijo:
                    tasa_aporte = p[prefijo + "aporte"] / 100
                    tasa_prima = p[prefijo + "prima"] / 100
                    tasa_comision = p[prefijo + "mixta"]/100 if row['Comisión AFP'] == "MIXTA" else p[prefijo + "flujo"]/100

                    aporte_afp = base_afp_onp * tasa_aporte
                    prima_afp = min(base_afp_onp, p['tope_afp']) * tasa_prima
                    comis_afp = base_afp_onp * tasa_comision
                    total_afp_ind = aporte_afp + prima_afp + comis_afp
                    if total_afp_ind > 0: desglose_descuentos[f'Aporte {sistema}'] = round(total_afp_ind, 2)

            total_pension = dscto_onp + aporte_afp + prima_afp + comis_afp

            # --- RENTA 5TA CATEGORÍA (Redondeo Entero PLAME) ---
            uit = p['uit']
            meses_restantes = 12 - mes_idx
            hist_q = historico_quinta.get(str(dni_trabajador), {})
            rem_previa_historica = hist_q.get('rem_previa', 0.0)
            retencion_previa_historica = hist_q.get('ret_previa', 0.0)
            # Proyección usa sueldo nominal completo: las ausencias son excepcionales
            base_quinta_proyeccion = base_quinta_mes
            if total_ausencias > 0 or ingreso_este_mes:
                base_quinta_proyeccion = round(base_quinta_mes + (sueldo_base_nominal - sueldo_computable) + conceptos_recuperados_5ta, 2)
            proyeccion_gratis = 0.0
            if mes_idx <= 6: proyeccion_gratis = base_quinta_proyeccion * 2 * 1.09
            elif mes_idx <= 11: proyeccion_gratis = base_quinta_proyeccion * 1 * 1.09
            proyeccion_sueldos_restantes = base_quinta_proyeccion * meses_restantes
            
            renta_bruta_anual = int(round(rem_previa_historica + base_quinta_mes + proyeccion_sueldos_restantes + proyeccion_gratis))
            renta_neta_anual = int(round(renta_bruta_anual - (7 * uit)))
            
            impuesto_anual = 0.0
            retencion_quinta = 0.0
            detalle_tramos = []
            divisor = 1
            if mes_idx in [1, 2, 3]: divisor = 12
            elif mes_idx == 4: divisor = 9
            elif mes_idx in [5, 6, 7]: divisor = 8
            elif mes_idx == 8: divisor = 5
            elif mes_idx in [9, 10, 11]: divisor = 4

            if renta_neta_anual > 0:
                renta_restante = renta_neta_anual
                tramos = [(5 * uit, 0.08), (15 * uit, 0.14), (15 * uit, 0.17), (10 * uit, 0.20), (float('inf'), 0.30)]
                for limite, tasa in tramos:
                    if renta_restante > 0:
                        monto_tramo = min(renta_restante, limite)
                        imp_tramo = monto_tramo * tasa
                        impuesto_anual += imp_tramo
                        detalle_tramos.append({"rango": f"Hasta {limite/uit} UIT", "tasa": f"{int(tasa*100)}%", "base": monto_tramo, "impuesto": imp_tramo})
                        renta_restante -= monto_tramo
                retencion_quinta = int(round(max(0.0, (impuesto_anual - retencion_previa_historica) / divisor)))
                if retencion_quinta > 0: desglose_descuentos['Retención 5ta Cat.'] = float(retencion_quinta)

            # --- APLICACIÓN DE AJUSTES DE AUDITORÍA (MANUALES) ---
            conceptos_manuales = json.loads(row.get('conceptos_json', '{}') or '{}')
            aj_afp    = float(conceptos_manuales.get('_ajuste_afp', 0.0))
            aj_quinta = float(conceptos_manuales.get('_ajuste_quinta', 0.0))
            aj_otros  = float(conceptos_manuales.get('_ajuste_otros', 0.0))

            if aj_afp != 0:
                desglose_descuentos['Ajuste AFP (Audit)'] = round(aj_afp, 2)
                descuentos_manuales += aj_afp
                obs_trab.append(f"Ajuste AFP: S/ {aj_afp:,.2f}")
            
            if aj_quinta != 0:
                desglose_descuentos['Ajuste 5ta Cat (Audit)'] = round(aj_quinta, 2)
                descuentos_manuales += aj_quinta
                obs_trab.append(f"Ajuste 5ta: S/ {aj_quinta:,.2f}")

            if aj_otros != 0:
                desglose_descuentos['Ajuste Varios (Audit)'] = round(aj_otros, 2)
                descuentos_manuales += aj_otros
                obs_trab.append(f"Ajuste Manual: S/ {aj_otros:,.2f}")

            # --- SEGURO SOCIAL (ESSALUD o SIS) Y NETO ---
            # Regla: EsSalud mínimo sobre RMV siempre que el trabajador tenga al
            # menos 1 día remunerado (trabajado o pagado).  Si el mes completo fue
            # suspensión sin goce de haber (días_remunerados == 0) → EsSalud = 0.
            seguro_social = str(row.get('Seguro Social', 'ESSALUD')).upper()
            _mes_completo_ssgh = (dias_remunerados == 0)

            if _mes_completo_ssgh:
                # Suspensión sin goce de haber todo el mes → sin aporte patronal
                aporte_essalud = 0.0
                etiqueta_seguro = ("SIS" if seguro_social == "SIS"
                                   else ("ESSALUD-EPS" if row.get('EPS', 'No') == "Sí"
                                         else "ESSALUD"))
            elif seguro_social == "SIS":
                aporte_essalud = 15.0  # Monto fijo SIS - Solo Micro Empresa
                etiqueta_seguro = "SIS"
            elif row.get('EPS', 'No') == "Sí":
                aporte_essalud = max(base_essalud, p['rmv']) * (p['tasa_eps'] / 100)
                etiqueta_seguro = "ESSALUD-EPS"
            else:
                aporte_essalud = max(base_essalud, p['rmv']) * (p['tasa_essalud'] / 100)
                etiqueta_seguro = "ESSALUD"
            
            neto_pagar = ingresos_totales - total_pension - retencion_quinta - descuentos_manuales

            # --- FILA DE LA SÁBANA CORPORATIVA ---
            resultados.append({
                "N°": index + 1,
                "DNI": dni_trabajador,
                "Apellidos y Nombres": nombres,
                "Sist. Pensión": sistema,
                "Seg. Social": etiqueta_seguro,
                "Sueldo Base": round(sueldo_computable, 2),
                "Asig. Fam.": round(monto_asig_fam, 2),
                "Otros Ingresos": round((pago_he_25 + pago_he_35 + monto_grati + otros_ingresos), 2),
                "TOTAL BRUTO": round(ingresos_totales, 2),
                "ONP (13%)": round(dscto_onp, 2),
                "AFP Aporte": round(aporte_afp, 2),
                "AFP Seguro": round(prima_afp, 2),
                "AFP Comis.": round(comis_afp, 2),
                "Ret. 5ta Cat.": float(retencion_quinta),
                "Dsctos/Faltas": round(descuentos_manuales, 2),
                "NETO A PAGAR": round(neto_pagar, 2),
                "Aporte Seg. Social": round(aporte_essalud, 2),
                # Alias para compatibilidad con boletas (leen 'EsSalud Patronal')
                "EsSalud Patronal": round(aporte_essalud, 2),
                # Datos bancarios para reporte de tesorería
                "Banco":     str(row.get('Banco', '') or ''),
                "N° Cuenta": str(row.get('Cuenta Bancaria', '') or ''),
                "CCI":       str(row.get('CCI', '') or ''),
                "Observaciones": " | ".join(obs_trab) if obs_trab else "",
            })

            auditoria_data[dni_trabajador] = {
                "nombres": nombres, "periodo": periodo_key,
                "dias": dias_laborados,               # días efectivamente laborados
                "dias_computables": dias_computables,  # base de proporcionalidad
                "observaciones": " | ".join(obs_trab),
                "rem_diaria": round(sueldo_base_nominal / 30.0, 2),
                "horas_ordinarias": horas_ordinarias,  # para .JOR de PLAME
                "suspensiones": susp_dict,             # para .SNL de PLAME
                "base_afp": round(base_afp_onp, 2),   # para AFPnet
                "seguro_social": etiqueta_seguro,
                "aporte_seg_social": round(aporte_essalud, 2),
                "ingresos": desglose_ingresos, "descuentos": desglose_descuentos,
                "totales": {"ingreso": ingresos_totales, "descuento": (total_pension + retencion_quinta + descuentos_manuales), "neto": neto_pagar},
                "quinta": {
                    "rem_previa": rem_previa_historica, "ret_previa": retencion_previa_historica,
                    "base_mes": base_quinta_mes, "meses_restantes": meses_restantes,
                    "proy_sueldo": proyeccion_sueldos_restantes, "proy_grati": proyeccion_gratis, 
                    "bruta_anual": renta_bruta_anual, "uit_valor": uit, "uit_7": 7 * uit, 
                    "neta_anual": renta_neta_anual, "detalle_tramos": detalle_tramos,
                    "imp_anual": impuesto_anual, "divisor": divisor, "retencion": retencion_quinta
                }
            }

        df_resultados = pd.DataFrame(resultados).fillna(0.0)
        
        # --- FILA DE TOTALES DINÁMICA ---
        cols_texto = {"N°", "DNI", "Apellidos y Nombres", "Sist. Pensión", "Seg. Social", "Banco", "N° Cuenta", "CCI", "Observaciones"}
        totales = {"N°": "", "DNI": "", "Apellidos y Nombres": "TOTALES", "Sist. Pensión": "", "Seg. Social": "", "Banco": "", "N° Cuenta": "", "CCI": "", "Observaciones": ""}
        for col in df_resultados.columns:
            if col not in cols_texto:
                totales[col] = df_resultados[col].sum()
            
        df_resultados = pd.concat([df_resultados, pd.DataFrame([totales])], ignore_index=True)
        st.session_state['res_planilla'] = df_resultados
        st.session_state['auditoria_data'] = auditoria_data

        # --- GUARDAR PLANILLA EN NEON (persistencia real) ---
        try:
            db2 = SessionLocal()
            _guardar_planilla(db2, empresa_id, periodo_key, df_resultados, auditoria_data)
            db2.close()
        except Exception as e:
            st.warning(f"Planilla calculada pero no se pudo guardar en la nube: {e}")

    # --- INTENTAR RECUPERAR PLANILLA GUARDADA EN NEON SI NO HAY EN SESSION ---
    if not st.session_state.get('ultima_planilla_calculada', False):
        try:
            db3 = SessionLocal()
            df_rec, aud_rec = _cargar_planilla_guardada(db3, empresa_id, periodo_key)
            db3.close()
            if df_rec is not None and not df_rec.empty:
                st.session_state['res_planilla'] = df_rec
                st.session_state['auditoria_data'] = aud_rec
                st.session_state['ultima_planilla_calculada'] = True
                if not es_cerrada:
                    st.info(f"📂 Planilla de **{periodo_key}** recuperada desde la nube.")
        except Exception:
            pass

    # --- RENDERIZADO VISUAL ---
    if st.session_state.get('ultima_planilla_calculada', False):
        df_resultados = st.session_state['res_planilla']
        auditoria_data = st.session_state.get('auditoria_data', {})

        if not es_cerrada:
            st.success("✅ Planilla generada con éxito.")

        st.markdown("### 📊 Matriz de Nómina")
        st.dataframe(df_resultados.iloc[:-1], use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("#### 📥 Exportación Corporativa (Planilla)")
        empresa_ruc_s = st.session_state.get('empresa_activa_ruc', '')
        empresa_reg_s = st.session_state.get('empresa_activa_regimen', '')
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            try:
                excel_file = generar_excel_sabana(df_resultados, empresa_nombre, periodo_key, empresa_ruc=empresa_ruc_s)
                st.download_button("📊 Descargar Sábana (.xlsx)", data=excel_file, file_name=f"PLANILLA_{periodo_key}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="dl_plan_xl")
            except Exception: pass
        with col_btn2:
            try:
                pdf_buffer = generar_pdf_sabana(df_resultados, empresa_nombre, periodo_key, empresa_ruc=empresa_ruc_s, empresa_regimen=empresa_reg_s)
                st.download_button("📄 Descargar Sábana y Resumen (PDF)", data=pdf_buffer, file_name=f"SABANA_{periodo_key}.pdf", mime="application/pdf", use_container_width=True, key="dl_plan_pdf")
            except Exception: pass

        st.markdown("---")
        st.markdown("### 🔒 Cierre de Planilla")

        rol_usuario = st.session_state.get('usuario_rol', 'analista')
        nombre_usuario = st.session_state.get('usuario_nombre', '')

        estado_actual = "ABIERTA"
        planilla_db_cierre = None
        try:
            db_c = SessionLocal()
            planilla_db_cierre = db_c.query(PlanillaMensual).filter_by(
                empresa_id=empresa_id, periodo_key=periodo_key
            ).first()
            db_c.close()
            if planilla_db_cierre:
                estado_actual = getattr(planilla_db_cierre, 'estado', 'ABIERTA') or 'ABIERTA'
        except Exception:
            pass

        if estado_actual == "CERRADA":
            cerrada_por  = getattr(planilla_db_cierre, 'cerrada_por', '') or ''
            fecha_cierre = getattr(planilla_db_cierre, 'fecha_cierre', None)
            fecha_str    = fecha_cierre.strftime("%d/%m/%Y %H:%M") if fecha_cierre else ''
            st.error(f"**PLANILLA CERRADA** — Responsable: {cerrada_por}  |  Fecha: {fecha_str}")
            if rol_usuario in ["supervisor", "admin"]:
                st.info("Como Administrador/Supervisor puede reabrir esta planilla para modificarla.")
                if st.button("🔓 Reabrir Planilla", use_container_width=False):
                    try:
                        db_up = SessionLocal()
                        p = db_up.query(PlanillaMensual).filter_by(
                            empresa_id=empresa_id, periodo_key=periodo_key
                        ).first()
                        p.estado = "ABIERTA"
                        p.cerrada_por = None
                        p.fecha_cierre = None
                        # Revertir cuotas de préstamos a PENDIENTE
                        _cuotas_rev = (
                            db_up.query(CuotaPrestamo)
                            .join(Prestamo)
                            .filter(
                                Prestamo.empresa_id == empresa_id,
                                CuotaPrestamo.periodo_key == periodo_key,
                                CuotaPrestamo.estado == 'PAGADA',
                            )
                            .all()
                        )
                        for _cr in _cuotas_rev:
                            _cr.estado = 'PENDIENTE'
                        db_up.commit()
                        db_up.close()
                        st.toast("Planilla REABIERTA para edición", icon="🔓")
                        st.rerun()
                    except Exception as e_re:
                        st.error(f"Error al reabrir: {e_re}")
            else:
                st.warning("Solo un **Supervisor** puede reabrir esta planilla.")
        else:
            st.info("La planilla está **ABIERTA**. Puede recalcularse hasta que sea cerrada.")
            if rol_usuario == "supervisor":
                with st.expander("Cerrar Planilla"):
                    st.warning("Al cerrar la planilla quedará bloqueada para el analista.")
                    if st.button("Confirmar Cierre de Planilla", type="primary"):
                        try:
                            db_up = SessionLocal()
                            p = db_up.query(PlanillaMensual).filter_by(
                                empresa_id=empresa_id, periodo_key=periodo_key
                            ).first()
                            p.estado = "CERRADA"
                            p.cerrada_por = nombre_usuario
                            p.fecha_cierre = datetime.now()
                            # Marcar cuotas del periodo como PAGADAS
                            _cuotas_pag = (
                                db_up.query(CuotaPrestamo)
                                .join(Prestamo)
                                .filter(
                                    Prestamo.empresa_id == empresa_id,
                                    CuotaPrestamo.periodo_key == periodo_key,
                                    CuotaPrestamo.estado == 'PENDIENTE',
                                )
                                .all()
                            )
                            for _cp in _cuotas_pag:
                                _cp.estado = 'PAGADA'
                            db_up.commit()
                            db_up.close()
                            db_up.commit()
                            db_up.close()
                            st.toast(f"Planilla {periodo_key} CERRADA exitosamente", icon="🔒")
                            st.rerun()
                        except Exception as e_cl:
                            st.error(f"Error al cerrar: {e_cl}")
            else:
                st.info("Solo un **Supervisor** puede cerrar la planilla.")


# ─── MOTOR DE HONORARIOS (4ta Categoría) ─────────────────────────────────────

def _render_honorarios_tab(empresa_id, empresa_nombre, periodo_key):
    """Motor de cálculo para Locadores de Servicio (4ta Categoría)."""
    from core.use_cases.calculo_honorarios import calcular_recibo_honorarios

    db = SessionLocal()
    try:
        # 1. Parámetros legales del periodo
        p = _cargar_parametros(db, empresa_id, periodo_key)
        if not p:
            st.error(f"🛑 No se han configurado los Parámetros Legales para **{periodo_key}**.")
            st.info("Vaya al módulo 'Parámetros Legales' y configure las tasas para este periodo.")
            return

        tasa_4ta = p.get('tasa_4ta', 8.0)
        tope_4ta = p.get('tope_4ta', 1500.0)

        # 2. Locadores activos
        mes_int  = int(periodo_key[:2])
        anio_int = int(periodo_key[3:])

        locadores_db = (
            db.query(Trabajador)
            .filter_by(empresa_id=empresa_id, situacion="ACTIVO", tipo_contrato="LOCADOR")
            .all()
        )

        # Filtrar locadores que ya iniciaron labores en o antes del periodo de cálculo
        locadores = [
            l for l in locadores_db
            if not (l.fecha_ingreso and (l.fecha_ingreso.year > anio_int or (l.fecha_ingreso.year == anio_int and l.fecha_ingreso.month > mes_int)))
        ]

        if not locadores:
            st.info("ℹ️ No hay Locadores de Servicio activos registrados en el Maestro de Personal.")
            return

        # 3. Variables del periodo para locadores
        variables_mes = (
            db.query(VariablesMes)
            .filter_by(empresa_id=empresa_id, periodo_key=periodo_key)
            .all()
        )
        vars_por_doc: dict = {}
        for v in variables_mes:
            dni = v.trabajador.num_doc
            conceptos_data = json.loads(v.conceptos_json or '{}')
            vars_por_doc[dni] = {
                'dias_no_prestados': getattr(v, 'dias_descuento_locador', 0) or 0,
                'otros_pagos':       float(conceptos_data.get('_otros_pagos_loc', 0.0) or 0.0),
                'otros_descuentos':  float(conceptos_data.get('_otros_descuentos_loc', 0.0) or 0.0),
            }

        # 4. Días del mes
        dias_del_mes = calendar.monthrange(anio_int, mes_int)[1]

    finally:
        db.close()

    # Verificar estado de cierre
    es_cerrada = False
    try:
        db_ck = SessionLocal()
        plan_ck = db_ck.query(PlanillaMensual).filter_by(
            empresa_id=empresa_id, periodo_key=periodo_key
        ).first()
        db_ck.close()
        if plan_ck and getattr(plan_ck, 'estado', 'ABIERTA') == 'CERRADA':
            es_cerrada = True
    except Exception:
        pass

    st.caption(f"Tasa retención 4ta Cat.: **{tasa_4ta}%** | Tope mínimo para retener: **S/ {tope_4ta:,.2f}**")

    if es_cerrada:
        st.error(f"🔒 Los honorarios del periodo **{periodo_key}** ya fueron CERRADOS.")
    elif st.button(f"🧮 Calcular Honorarios - {periodo_key}", type="primary", use_container_width=True):
        resultados_loc = []
        
        # Precargar cuotas de préstamos para locadores
        cuotas_loc = {}
        try:
            db_cl = SessionLocal()
            _c_loc = db_cl.query(CuotaPrestamo).join(Prestamo).filter(
                Prestamo.empresa_id == empresa_id,
                CuotaPrestamo.periodo_key == periodo_key,
                CuotaPrestamo.estado == 'PENDIENTE'
            ).all()
            for _cloc in _c_loc:
                cuotas_loc.setdefault(_cloc.prestamo.trabajador.num_doc, []).append(_cloc)
            db_cl.close()
        except Exception:
            pass

        # Precargar notas para locadores
        notas_loc_map = {}
        try:
            db_nl = SessionLocal()
            _v_nl = db_nl.query(VariablesMes).filter_by(empresa_id=empresa_id, periodo_key=periodo_key).all()
            for _vnl in _v_nl:
                notas_loc_map[_vnl.trabajador.num_doc] = getattr(_vnl, 'notas_gestion', '') or ''
            db_nl.close()
        except: pass

        for loc in locadores:
            dni = loc.num_doc
            vars_loc = vars_por_doc.get(dni, {})
            
            # Sumar cuotas de préstamos a otros descuentos del locador
            monto_cuotas_loc = sum(float(c.monto) for c in cuotas_loc.get(dni, []))
            if monto_cuotas_loc > 0:
                vars_loc['otros_descuentos'] = vars_loc.get('otros_descuentos', 0.0) + monto_cuotas_loc
                cuotas_desc = [f"{c.prestamo.concepto} (Cuota {c.numero_cuota})" for c in cuotas_loc.get(dni, [])]
                obs_p = f"Dscto. Préstamos: {', '.join(cuotas_desc)}"
            else:
                obs_p = ""

            resultado = calcular_recibo_honorarios(
                loc, vars_loc, dias_del_mes,
                tasa_4ta=tasa_4ta, tope_4ta=tope_4ta,
                anio_calc=anio_int, mes_calc=mes_int,
            )
            
            if obs_p:
                resultado['observaciones'] = f"{resultado['observaciones']} | {obs_p}" if resultado['observaciones'] else obs_p
            
            # Integrar Nota de Gestión Manual para locadores
            nota_m_loc = notas_loc_map.get(dni, "")
            if nota_m_loc:
                resultado['observaciones'] = f"{resultado['observaciones']} | NOTA: {nota_m_loc}" if resultado['observaciones'] else f"NOTA: {nota_m_loc}"
            
            # Agregar monto de descuento por días no prestados para Tesorería
            if resultado['dias_no_prestados'] > 0:
                desc_dias = f"Días no prestados: {resultado['dias_no_prestados']} (Desc: S/ {resultado['monto_descuento']:,.2f})"
                resultado['observaciones'] = f"{resultado['observaciones']} | {desc_dias}" if resultado['observaciones'] else desc_dias

            resultados_loc.append({
                "DNI":                 dni,
                "Locador":             loc.nombres,
                "Honorario Base":      resultado['honorario_base'],
                "Días Laborados":      resultado['dias_laborados'],
                "Días no Prestados":   resultado['dias_no_prestados'],
                "Descuento Días":      resultado['monto_descuento'],
                "Otros Pagos":         resultado['otros_pagos'],
                "Pago Bruto":          resultado['pago_bruto'],
                "Retención 4ta (8%)":  resultado['retencion_4ta'],
                "Otros Descuentos":    resultado['otros_descuentos'],
                "NETO A PAGAR":        resultado['neto_a_pagar'],
                "Banco":               getattr(loc, 'banco', '') or '',
                "N° Cuenta":           getattr(loc, 'cuenta_bancaria', '') or '',
                "CCI":                 getattr(loc, 'cci', '') or '',
                "Observaciones":       resultado['observaciones'],
            })
        st.session_state[f'res_honorarios_{periodo_key}'] = pd.DataFrame(resultados_loc)

    key_res = f'res_honorarios_{periodo_key}'
    if st.session_state.get(key_res) is not None and not st.session_state[key_res].empty:
        df_loc = st.session_state[key_res]
        st.success("✅ Valorización de Honorarios generada.")
        st.dataframe(df_loc, use_container_width=True, hide_index=True)

        c1, c2, c3 = st.columns(3)
        c1.metric("Total Pago Bruto",      f"S/ {df_loc['Pago Bruto'].sum():,.2f}")
        c2.metric("Total Retención 4ta",   f"S/ {df_loc['Retención 4ta (8%)'].sum():,.2f}")
        c3.metric("Total Neto a Pagar",    f"S/ {df_loc['NETO A PAGAR'].sum():,.2f}")

        st.markdown("---")
        st.markdown("#### 📥 Exportación Corporativa (Locadores)")
        col_h1, col_h2 = st.columns(2)
        empresa_ruc_h = st.session_state.get('empresa_activa_ruc', '')
        empresa_reg_h = st.session_state.get('empresa_activa_regimen', '')
        with col_h1:
            buf_xls = generar_excel_honorarios(df_loc, empresa_nombre, periodo_key, empresa_ruc=empresa_ruc_h)
            st.download_button(
                "📊 Descargar Valorización (.xlsx)",
                data=buf_xls,
                file_name=f"HONORARIOS_{periodo_key}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_h2:
            pdf_hon = generar_pdf_honorarios(df_loc, empresa_nombre, periodo_key, empresa_ruc=empresa_ruc_h, empresa_regimen=empresa_reg_h)
            st.download_button(
                "📄 Descargar Valorización (PDF)",
                data=pdf_hon,
                file_name=f"HONORARIOS_{periodo_key}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
    else:
        if not es_cerrada:
            st.info("Presione el botón para calcular los honorarios del periodo.")


# ─── RENDER PRINCIPAL ─────────────────────────────────────────────────────────

def render():
    st.title("⚙️ Ejecución de Planilla Mensual")
    st.markdown("---")

    empresa_id     = st.session_state.get('empresa_activa_id')
    empresa_nombre = st.session_state.get('empresa_activa_nombre')

    col_m, col_a = st.columns([2, 1])
    mes_seleccionado  = col_m.selectbox("Mes de Cálculo", MESES, key="calc_mes")
    anio_seleccionado = col_a.selectbox("Año de Cálculo", [2025, 2026, 2027, 2028], index=1, key="calc_anio")
    periodo_key = f"{mes_seleccionado[:2]}-{anio_seleccionado}"
    mes_idx = MESES.index(mes_seleccionado) + 1

    tab_plan, tab_hon = st.tabs(["📋 1. Planilla (5ta Categoría)", "🧾 2. Honorarios (4ta Categoría)"])

    with tab_plan:
        _render_planilla_tab(empresa_id, empresa_nombre, mes_seleccionado, anio_seleccionado, periodo_key, mes_idx)

    with tab_hon:
        _render_honorarios_tab(empresa_id, empresa_nombre, periodo_key)


    # ── SECCIÓN GLOBAL DE TESORERÍA ──────────────────────────────────────────
    df_plan_glob = st.session_state.get('res_planilla', pd.DataFrame())
    df_loc_glob  = st.session_state.get(f'res_honorarios_{periodo_key}', pd.DataFrame())
    aud_glob  = st.session_state.get('auditoria_data', {})

    if not df_plan_glob.empty or not df_loc_glob.empty:
        st.markdown("---")
        st.subheader("🏦 Gestión de Tesorería")

        try:
            _db_chk = SessionLocal()
            _n_loc_glob = _db_chk.query(Trabajador).filter_by(empresa_id=empresa_id, situacion='ACTIVO', tipo_contrato='LOCADOR').count()
            _db_chk.close()
        except:
            _n_loc_glob = 0
        
        if _n_loc_glob > 0 and df_loc_glob.empty:
            st.warning("⚠️ Locadores detectados. Calcule Honorarios en la pestaña '🧾 2' para un reporte completo.")
        
        try:
            buf_teso_f = generar_pdf_tesoreria(
                df_planilla=df_plan_glob if not df_plan_glob.empty else None,
                df_loc=df_loc_glob if not df_loc_glob.empty else None,
                empresa_nombre=empresa_nombre,
                periodo_key=periodo_key,
                auditoria_data=aud_glob,
                empresa_ruc=st.session_state.get('empresa_activa_ruc', ''),
            )
            st.download_button(
                "🏦 Descargar Reporte de Tesorería (PDF)",
                data=buf_teso_f,
                file_name=f"TESORERIA_{periodo_key}.pdf",
                mime="application/pdf",
                use_container_width=True,
                type="primary",
                key="btn_teso_global_v_final"
            )
        except Exception: pass

        st.markdown("---")
        with st.expander("🔍 Panel de Auditoría Tributaria y Liquidaciones", expanded=False):
            if not aud_glob:
                st.info("No hay datos de auditoría calculados.")
            else:
                opciones_trab = [f"{dni} - {info['nombres']}" for dni, info in aud_glob.items()]
                trabajador_sel = st.selectbox("Seleccione un trabajador para ver su detalle legal:", opciones_trab, label_visibility="collapsed")
                
                if trabajador_sel:
                    dni_sel = trabajador_sel.split(" - ")[0]
                    data = aud_glob[dni_sel]
                    q = data['quinta']
                    t_audit1, t_audit2 = st.tabs(["💰 Boleta Mensual", "🏛️ Certificado de 5ta Categoría"])
                    with t_audit1:
                        c_a1, c_a2 = st.columns(2)
                        with c_a1:
                            for k, v in data['ingresos'].items(): st.markdown(f"- **{k}:** S/ {v:,.2f}")
                            st.success(f"**Total Ingresos: S/ {data['totales']['ingreso']:,.2f}**")
                        with c_a2:
                            for k, v in data['descuentos'].items(): st.markdown(f"- **{k}:** S/ {v:,.2f}")
                            st.error(f"**Total Descuentos: S/ {data['totales']['descuento']:,.2f}**")
                    with t_audit2:
                        if q['neta_anual'] <= 0: st.success("Este trabajador NO supera las 7 UIT anuales.")
                        else:
                            pdf_5ta = generar_pdf_quinta(q, empresa_nombre, periodo_key, data['nombres'])
                            st.download_button("📄 Descargar Certificado de 5ta Categoría (PDF)", data=pdf_5ta, file_name=f"QUINTA_{dni_sel}_{periodo_key}.pdf", mime="application/pdf")

    # ── REPORTE COMBINADO (Planilla + Locadores) ───────────────────────────────
    df_plan_comb = st.session_state.get('res_planilla', pd.DataFrame())
    df_loc_comb  = st.session_state.get(f'res_honorarios_{periodo_key}', pd.DataFrame())

    if not df_plan_comb.empty:
        # Gate del reporte combinado: misma lógica que en _render_planilla_tab
        try:
            _db_gc = SessionLocal()
            _n_loc_gc = _db_gc.query(Trabajador).filter_by(
                empresa_id=empresa_id, situacion='ACTIVO', tipo_contrato='LOCADOR'
            ).count()
            _db_gc.close()
        except Exception:
            _n_loc_gc = 0
        _gate_comb = (_n_loc_gc == 0) or (not df_loc_comb.empty)

        st.markdown("---")
        with st.expander("📋 Reporte Consolidado de Costo Laboral (Planilla + Locadores)", expanded=False):
            st.markdown(
                "Genera un único documento PDF/Excel que combina la sábana de planilla y la "
                "valorización de locadores, más un resumen del costo laboral total de la empresa."
            )
            if not _gate_comb:
                st.warning(
                    "⚠️ Esta empresa tiene **Locadores de Servicio activos**. "
                    "Calcule primero los **Honorarios** en la pestaña '🧾 2. Honorarios (4ta Categoría)' "
                    "para que el reporte incluya la información completa."
                )
            elif df_loc_comb.empty:
                st.info("ℹ️ No hay datos de honorarios calculados para este periodo. El reporte combinado incluirá solo la planilla.")

            if _gate_comb:
                empresa_ruc_c = st.session_state.get('empresa_activa_ruc', '')
                empresa_reg_c = st.session_state.get('empresa_activa_regimen', '')
                col_comb1, col_comb2 = st.columns(2)
                with col_comb1:
                    pdf_comb = generar_pdf_combinado(
                        df_plan_comb,
                        df_loc_comb if not df_loc_comb.empty else None,
                        empresa_nombre, periodo_key,
                        empresa_ruc=empresa_ruc_c, empresa_regimen=empresa_reg_c
                    )
                    st.download_button(
                        "📄 Descargar Reporte Combinado (PDF)",
                        data=pdf_comb,
                        file_name=f"COSTO_LABORAL_{periodo_key}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        type="primary",
                    )
                with col_comb2:
                    buf_comb_xl = io.BytesIO()
                    with pd.ExcelWriter(buf_comb_xl, engine='openpyxl') as writer:
                        df_plan_comb.to_excel(writer, sheet_name=f'Planilla_{periodo_key[:2]}', index=False)
                        if not df_loc_comb.empty:
                            # Excluir observaciones y datos bancarios de la hoja de locadores en Excel combinado
                            cols_xl_loc = [c for c in df_loc_comb.columns if c not in ["Observaciones", "Banco", "N° Cuenta", "CCI"]]
                            df_loc_comb[cols_xl_loc].to_excel(writer, sheet_name=f'Honorarios_{periodo_key[:2]}', index=False)
                    buf_comb_xl.seek(0)
                    st.download_button(
                        "📊 Descargar Reporte Combinado (.xlsx)",
                        data=buf_comb_xl,
                        file_name=f"COSTO_LABORAL_{periodo_key}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

