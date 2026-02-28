"""
Módulo de Préstamos y Descuentos Programados.

Permite otorgar préstamos/descuentos a trabajadores con cuotas mensuales.
Las cuotas se aplican automáticamente en el motor de cálculo de planilla.
"""

import streamlit as st
import pandas as pd
import io
from datetime import date, datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import portrait, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from infrastructure.database.connection import SessionLocal
from infrastructure.database.models import Trabajador, Prestamo, CuotaPrestamo


# ─── HELPERS DE REPORTES ───────────────────────────────────────────────────────

def generar_excel_cronograma(dp, empresa_nombre):
    buffer = io.BytesIO()
    df = pd.DataFrame(dp['cuotas'])
    df = df[['numero_cuota', 'periodo_key', 'monto', 'estado']]
    df.columns = ['N° Cuota', 'Periodo', 'Monto (S/)', 'Estado']
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=6)
        ws = writer.sheets['Sheet1']
        ws['A1'] = empresa_nombre.upper()
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"CRONOGRAMA DE PAGOS - {dp['concepto'].upper()}"
        ws['A3'] = f"TRABAJADOR: {dp['trabajador']} ({dp['dni']})"
        ws['A4'] = f"MONTO TOTAL: S/ {dp['monto_total']:,.2f}"
    buffer.seek(0)
    return buffer

def generar_pdf_cronograma(dp, empresa_nombre):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=portrait(letter))
    elements = []
    styles = getSampleStyleSheet()
    
    C_NAVY = colors.HexColor("#0F2744")
    st_h = ParagraphStyle('H', fontName="Helvetica-Bold", fontSize=12, textColor=C_NAVY)
    
    elements.append(Paragraph(empresa_nombre.upper(), st_h))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>Cronograma de:</b> {dp['concepto']}", styles['Normal']))
    elements.append(Paragraph(f"<b>Trabajador:</b> {dp['trabajador']} ({dp['dni']})", styles['Normal']))
    elements.append(Paragraph(f"<b>Monto Total:</b> S/ {dp['monto_total']:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 15))
    
    data = [["N° CUOTA", "PERIODO", "MONTO (S/)", "ESTADO"]]
    for c in dp['cuotas']:
        data.append([c['numero_cuota'], c['periodo_key'], f"{c['monto']:,.2f}", c['estado']])
        
    t = Table(data, colWidths=[80, 120, 120, 120])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), C_NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer

# ─── HELPER ────────────────────────────────────────────────────────────────────

def _sumar_un_mes(periodo_key: str) -> str:
    """Incrementa en un mes el periodo_key con formato MM-YYYY."""
    mes, anio = int(periodo_key[:2]), int(periodo_key[3:])
    mes += 1
    if mes > 12:
        mes = 1
        anio += 1
    return f"{mes:02d}-{anio}"


# ─── RENDER PRINCIPAL ──────────────────────────────────────────────────────────

def render():
    empresa_id     = st.session_state.get('empresa_activa_id')
    empresa_nombre = st.session_state.get('empresa_activa_nombre', '')

    if not empresa_id:
        st.warning("⚠️ Seleccione una empresa para gestionar préstamos.")
        return

    st.title("💳 Préstamos y Descuentos Programados")
    st.markdown("---")

    tab_otorgar, tab_cronogramas = st.tabs(["➕ Otorgar Descuento", "📋 Cronogramas Activos"])

    # ── TAB 1: OTORGAR DESCUENTO ───────────────────────────────────────────────
    with tab_otorgar:
        st.markdown("#### Registrar nuevo préstamo o descuento programado")

        db = SessionLocal()
        try:
            trabajadores = (
                db.query(Trabajador)
                .filter_by(empresa_id=empresa_id, situacion="ACTIVO", tipo_contrato="PLANILLA")
                .order_by(Trabajador.nombres)
                .all()
            )
        finally:
            db.close()

        if not trabajadores:
            st.info("No hay trabajadores activos en planilla para esta empresa.")
            return

        opciones_trab = {f"{t.nombres} — DNI {t.num_doc}": t.id for t in trabajadores}

        with st.form("form_prestamo", clear_on_submit=True):
            trab_sel = st.selectbox("Trabajador", list(opciones_trab.keys()))

            _CONCEPTOS_BASE = [
                "Préstamo Personal",
                "Adelanto de Sueldo",
                "Faltante de Caja",
                "Pérdida de Activo",
                "Otro (Especificar)",
            ]
            concepto_sel = st.selectbox("Concepto del descuento", _CONCEPTOS_BASE)
            concepto_personalizado = ""
            if concepto_sel == "Otro (Especificar)":
                concepto_personalizado = st.text_input("Especifique el concepto")

            col1, col2 = st.columns(2)
            monto_total   = col1.number_input("Monto total (S/)", min_value=1.0, step=50.0, format="%.2f")
            n_cuotas      = col2.number_input("N° de cuotas", min_value=1, max_value=60, step=1, value=1)

            # Mes de inicio en formato MM-YYYY
            hoy = date.today()
            mes_inicio_default = f"{hoy.month:02d}-{hoy.year}"
            mes_inicio = st.text_input(
                "Mes de inicio (MM-YYYY)",
                value=mes_inicio_default,
                help="Formato: MM-YYYY, ej: 03-2026"
            )

            submitted = st.form_submit_button("💾 Guardar Descuento Programado", type="primary")

        if submitted:
            # Validar mes_inicio
            try:
                int(mes_inicio[:2]); int(mes_inicio[3:])
                assert len(mes_inicio) == 7 and mes_inicio[2] == '-'
            except Exception:
                st.error("El mes de inicio debe tener formato MM-YYYY (ej: 03-2026).")
                return

            concepto_final = concepto_personalizado.strip() if concepto_sel == "Otro (Especificar)" else concepto_sel
            if not concepto_final:
                st.error("Debe especificar el concepto del descuento.")
                return

            trabajador_id = opciones_trab[trab_sel]
            monto_cuota   = round(monto_total / n_cuotas, 2)
            # Ajuste de redondeo en la última cuota
            diferencia    = round(monto_total - monto_cuota * n_cuotas, 2)

            db = SessionLocal()
            try:
                nuevo_prestamo = Prestamo(
                    empresa_id         = empresa_id,
                    trabajador_id      = trabajador_id,
                    concepto           = concepto_final,
                    monto_total        = monto_total,
                    numero_cuotas      = n_cuotas,
                    fecha_otorgamiento = hoy,
                    estado             = "ACTIVO",
                )
                db.add(nuevo_prestamo)
                db.flush()  # Obtener nuevo_prestamo.id antes del commit

                periodo_iter = mes_inicio
                for i in range(n_cuotas):
                    monto_c = monto_cuota + (diferencia if i == n_cuotas - 1 else 0.0)
                    cuota = CuotaPrestamo(
                        prestamo_id  = nuevo_prestamo.id,
                        numero_cuota = i + 1,
                        periodo_key  = periodo_iter,
                        monto        = round(monto_c, 2),
                        estado       = "PENDIENTE",
                    )
                    db.add(cuota)
                    periodo_iter = _sumar_un_mes(periodo_iter)

                db.commit()
                st.success(
                    f"✅ **{concepto_final}** registrado correctamente. "
                    f"{n_cuotas} cuota(s) de S/ {monto_cuota:,.2f} a partir de {mes_inicio}."
                )
            except Exception as e:
                db.rollback()
                st.error(f"Error al guardar: {e}")
            finally:
                db.close()

    # ── TAB 2: CRONOGRAMAS ACTIVOS ─────────────────────────────────────────────
    with tab_cronogramas:
        st.markdown("#### Préstamos y descuentos vigentes")

        rol_usuario = st.session_state.get('usuario_rol', 'analista')

        db = SessionLocal()
        try:
            prestamos_list = (
                db.query(Prestamo)
                .filter_by(empresa_id=empresa_id, estado="ACTIVO")
                .join(Prestamo.trabajador)
                .order_by(Trabajador.nombres)
                .all()
            )
            # Cargar cuotas eager (dentro de la sesión)
            datos_prestamos = []
            for pr in prestamos_list:
                cuotas_data = []
                for c in sorted(pr.cuotas, key=lambda x: x.numero_cuota):
                    cuotas_data.append({
                        'id':            c.id,
                        'numero_cuota':  c.numero_cuota,
                        'periodo_key':   c.periodo_key,
                        'monto':         c.monto,
                        'estado':        c.estado,
                    })
                pagado    = sum(c['monto'] for c in cuotas_data if c['estado'] == 'PAGADA')
                pendiente = sum(c['monto'] for c in cuotas_data if c['estado'] == 'PENDIENTE')
                datos_prestamos.append({
                    'id':            pr.id,
                    'trabajador':    pr.trabajador.nombres,
                    'dni':           pr.trabajador.num_doc,
                    'concepto':      pr.concepto,
                    'monto_total':   pr.monto_total,
                    'numero_cuotas': pr.numero_cuotas,
                    'pagado':        pagado,
                    'pendiente':     pendiente,
                    'cuotas':        cuotas_data,
                })
        finally:
            db.close()

        if not datos_prestamos:
            st.info("No hay préstamos o descuentos activos para esta empresa.")
            return

        for dp in datos_prestamos:
            titulo_exp = (
                f"**{dp['trabajador']}** (DNI {dp['dni']}) — "
                f"{dp['concepto']} | Total: S/ {dp['monto_total']:,.2f} | "
                f"Pagado: S/ {dp['pagado']:,.2f} | Saldo: S/ {dp['pendiente']:,.2f}"
            )
            with st.expander(titulo_exp, expanded=False):
                col_r, col_c = st.columns([3, 1])
                col_r.markdown(
                    f"**Concepto:** {dp['concepto']}  \n"
                    f"**Total:** S/ {dp['monto_total']:,.2f}  \n"
                    f"**Cuotas:** {dp['numero_cuotas']}  \n"
                    f"**Pagado:** S/ {dp['pagado']:,.2f}  \n"
                    f"**Saldo Pendiente:** S/ {dp['pendiente']:,.2f}"
                )

                # Botones de descarga
                buf_xl = generar_excel_cronograma(dp, empresa_nombre)
                col_c.download_button("📊 Excel", data=buf_xl, file_name=f"Cronograma_{dp['dni']}.xlsx", use_container_width=True)
                
                buf_pdf = generar_pdf_cronograma(dp, empresa_nombre)
                col_c.download_button("📄 PDF", data=buf_pdf, file_name=f"Cronograma_{dp['dni']}.pdf", use_container_width=True)

                # Botón para cancelar el préstamo completo (solo supervisores)
                if rol_usuario == "supervisor" and col_c.button(
                    "🗑️ Cancelar",
                    key=f"cancel_prest_{dp['id']}",
                    use_container_width=True
                ):
                    db2 = SessionLocal()
                    try:
                        pr_obj = db2.query(Prestamo).get(dp['id'])
                        if pr_obj:
                            pr_obj.estado = "CANCELADO"
                            db2.commit()
                            st.success("Préstamo cancelado.")
                            st.rerun()
                    except Exception as e2:
                        db2.rollback()
                        st.error(f"Error: {e2}")
                    finally:
                        db2.close()

                st.markdown("---")
                # Tabla de cuotas
                _COL_N  = "N° Cuota"
                _COL_P  = "Periodo"
                _COL_M  = "Monto (S/)"
                _COL_E  = "Estado"
                _COL_AC = "Acción"

                header_cols = st.columns([1, 2, 2, 2, 2])
                for hdr, col in zip([_COL_N, _COL_P, _COL_M, _COL_E, _COL_AC], header_cols):
                    col.markdown(f"**{hdr}**")

                for cuota in dp['cuotas']:
                    c1, c2, c3, c4, c5 = st.columns([1, 2, 2, 2, 2])
                    c1.write(str(cuota['numero_cuota']))
                    c2.write(cuota['periodo_key'])
                    c3.write(f"S/ {cuota['monto']:,.2f}")
                    estado_badge = (
                        f"✅ Pagada" if cuota['estado'] == 'PAGADA'
                        else f"🕐 Pendiente"
                    )
                    c4.write(estado_badge)

                    # Aplazamiento Dominó — solo supervisor + cuota PENDIENTE
                    if (
                        rol_usuario == "supervisor"
                        and cuota['estado'] == 'PENDIENTE'
                    ):
                        if c5.button(
                            "⏭️ Aplazar 1 mes",
                            key=f"aplazar_{cuota['id']}",
                            use_container_width=True
                        ):
                            db3 = SessionLocal()
                            try:
                                # Dominó: actualizar ESTA cuota y todas las siguientes del mismo préstamo
                                cuotas_afectadas = (
                                    db3.query(CuotaPrestamo)
                                    .filter(
                                        CuotaPrestamo.prestamo_id == dp['id'],
                                        CuotaPrestamo.numero_cuota >= cuota['numero_cuota'],
                                        CuotaPrestamo.estado == 'PENDIENTE',
                                    )
                                    .all()
                                )
                                for c_afect in cuotas_afectadas:
                                    c_afect.periodo_key = _sumar_un_mes(c_afect.periodo_key)
                                db3.commit()
                                st.success(
                                    f"Cuota {cuota['numero_cuota']} y subsiguientes aplazadas 1 mes."
                                )
                                st.rerun()
                            except Exception as e3:
                                db3.rollback()
                                st.error(f"Error al aplazar: {e3}")
                            finally:
                                db3.close()
                    else:
                        c5.write("—")
