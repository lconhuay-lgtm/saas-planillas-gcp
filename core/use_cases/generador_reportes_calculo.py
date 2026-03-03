"""
Módulo de generación de reportes de cálculo de planillas.

Contiene todas las funciones de exportación (PDF y Excel) para:
  - Sábana de planilla (5ta categoría)
  - Valorización de locadores (4ta categoría)
  - Reporte combinado
  - Reporte de tesorería
  - Certificados de 5ta categoría
  - Reportes personalizados

Estas funciones son puras respecto a la UI: reciben DataFrames y retornan
buffers de bytes listos para descargar. No dependen de Streamlit.
"""
import io
import calendar
from datetime import datetime

import pandas as pd

# Excel
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, portrait, letter, legal
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

_MESES_ES_CALC = {
    "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
    "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
}

def _periodo_legible_calc(periodo_key: str) -> str:
    """'02-2026' → 'Febrero - 2026'"""
    partes = periodo_key.split("-")
    if len(partes) == 2:
        return f"{_MESES_ES_CALC.get(partes[0], partes[0])} - {partes[1]}"
    return periodo_key

# Columnas internas que NO deben aparecer en los reportes (aliases/duplicados)
_COLS_OCULTAS_SABANA = {"EsSalud Patronal"}

# Paleta corporativa de colores compartida por todos los generadores de PDF
C_NAVY  = colors.HexColor("#0F2744")
C_STEEL = colors.HexColor("#1E4D8C")
C_GOLD  = colors.HexColor("#C9A84C")
C_LIGHT = colors.HexColor("#F0F4F9")
C_GRAY  = colors.HexColor("#64748B")


def _make_canvas_header(empresa_nombre, empresa_ruc, linea_titulo, fecha_emision):
    """Devuelve el callback draw_header para doc.build() con membrete corporativo fijo."""
    def _dh(canvas_obj, doc_obj):
        canvas_obj.saveState()
        page_w, page_h = landscape(legal)
        x0, x1 = 12, page_w - 12
        offset_ruc = 14 if empresa_ruc else 0

        canvas_obj.setFont("Helvetica-Bold", 13)
        canvas_obj.setFillColor(C_NAVY)
        canvas_obj.drawString(x0, page_h - 22, empresa_nombre.upper())

        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.setFillColor(colors.HexColor("#64748B"))
        canvas_obj.drawRightString(x1, page_h - 22, f"Pág. {doc_obj.page}")

        if empresa_ruc:
            canvas_obj.setFont("Helvetica", 9)
            canvas_obj.setFillColor(colors.HexColor("#64748B"))
            canvas_obj.drawString(x0, page_h - 36, f"RUC: {empresa_ruc}")

        canvas_obj.setFont("Helvetica-Bold", 9)
        canvas_obj.setFillColor(C_STEEL)
        canvas_obj.drawString(x0, page_h - 36 - offset_ruc, linea_titulo)

        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.setFillColor(colors.HexColor("#64748B"))
        canvas_obj.drawString(x0, page_h - 50 - offset_ruc, f"Emitido el {fecha_emision}")

        canvas_obj.setStrokeColor(C_NAVY)
        canvas_obj.setLineWidth(1.2)
        canvas_obj.line(x0, page_h - 63 - offset_ruc, x1, page_h - 63 - offset_ruc)
        canvas_obj.setStrokeColor(C_GOLD)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(x0, page_h - 67 - offset_ruc, x1, page_h - 67 - offset_ruc)

        canvas_obj.restoreState()
    return _dh


def generar_excel_sabana(df, empresa_nombre, periodo, empresa_ruc=""):
    """Genera un archivo Excel nativo (.xlsx) profesional con colores corporativos"""
    periodo_texto = _periodo_legible_calc(periodo)
    # Excluir columna duplicada
    cols_mostrar = [c for c in df.columns if c not in _COLS_OCULTAS_SABANA]
    df_export = df[cols_mostrar]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name=f'Planilla_{periodo[:2]}', index=False, startrow=5)
        ws = writer.sheets[f'Planilla_{periodo[:2]}']

        # 1. Títulos de Cabecera
        ws['A1'] = empresa_nombre
        ws['A1'].font = Font(size=16, bold=True, color="0F2744")
        ws['A2'] = empresa_ruc and f"RUC: {empresa_ruc}" or ""
        ws['A2'].font = Font(size=10, color="64748B")
        ws['A3'] = f"PLANILLA DE REMUNERACIONES — PERIODO: {periodo_texto}"
        ws['A3'].font = Font(size=11, bold=True, color="1E4D8C")
        ws['A4'] = f"Fecha de Cálculo: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A4'].font = Font(size=10, italic=True, color="7F8C8D")
        
        # 2. Estilos Corporativos
        fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        
        # 3. Aplicar formato a las celdas de la tabla (datos desde fila 6)
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_thin
                if cell.row == 6:  # Fila de nombres de columnas
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = align_center
                elif cell.row == ws.max_row:  # Última fila (Totales)
                    cell.fill = fill_total
                    cell.font = Font(bold=True)
                    
        # 4. Auto-ajustar ancho de columnas (máximo 25 caracteres para no hacerlo gigante)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 25)
            
    buffer.seek(0)
    return buffer

def generar_pdf_sabana(df, empresa_nombre, periodo, empresa_ruc="", empresa_regimen=""):
    """Genera la sábana principal de planilla — ajustada a hoja, sin overflow."""
    periodo_texto = _periodo_legible_calc(periodo)

    # 1. Excluir columnas internas/duplicadas y datos bancarios (exclusivos del Reporte Tesorería)
    _OCULTAR_PDF = _COLS_OCULTAS_SABANA | {"Banco", "N° Cuenta", "CCI"}
    cols_work = [c for c in df.columns if c not in _OCULTAR_PDF]
    df = df[cols_work].copy()

    # 2. Excluir columnas numéricas donde TODOS los trabajadores tienen 0
    #    (ej. "Otros Ingresos" si nadie tiene bonos extra). Se compara solo sobre filas de datos, no la fila TOTALES.
    _COLS_IDENTIDAD = {"N°", "DNI", "Apellidos y Nombres", "Sist. Pensión", "Seg. Social"}
    if 'Apellidos y Nombres' in df.columns:
        df_datos = df[df['Apellidos y Nombres'] != 'TOTALES']
    else:
        df_datos = df.iloc[:-1]
    cols_mostrar = [
        c for c in df.columns
        if c in _COLS_IDENTIDAD
        or (c not in df_datos.columns)
        or pd.to_numeric(df_datos[c], errors='coerce').fillna(0).sum() != 0
    ]
    df = df[cols_mostrar]

    # ── Página landscape legal (1008 × 612 pt), márgenes 12 ──
    W_PAGE = 1008 - 24   # 984 pt disponibles
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=15, bottomMargin=15
    )
    elements = []
    styles = getSampleStyleSheet()

    st_title = ParagraphStyle('T', fontName="Helvetica-Bold", fontSize=14, textColor=C_NAVY, spaceAfter=6)
    st_sub   = ParagraphStyle('S', fontName="Helvetica",      fontSize=9,  textColor=C_GRAY, spaceAfter=1)
    st_head  = ParagraphStyle('H', fontName="Helvetica-Bold", fontSize=10, textColor=C_STEEL, spaceAfter=8, spaceBefore=4)

    fecha_calc = datetime.now().strftime("%d/%m/%Y %H:%M")
    ruc_line = f"  |  RUC: {empresa_ruc}" if empresa_ruc else ""
    reg_line = f"  |  {empresa_regimen}" if empresa_regimen else ""
    elements.append(Paragraph(empresa_nombre + ruc_line, st_title))
    elements.append(Paragraph(
        f"PLANILLA DE REMUNERACIONES  ·  PERIODO: {periodo_texto}{reg_line}", st_head
    ))
    elements.append(Paragraph(f"Fecha de cálculo: {fecha_calc}", st_sub))
    elements.append(Spacer(1, 8))

    # ── Anchos de columna proporcionales (suman W_PAGE) ──
    # Orden de columnas esperado tras excluir EsSalud Patronal:
    col_widths_map = {
        "N°": 18, "DNI": 52, "Apellidos y Nombres": 108,
        "Sist. Pensión": 58, "Seg. Social": 50,
        "Sueldo Base": 52, "Asig. Fam.": 38, "Otros Ingresos": 52,
        "TOTAL BRUTO": 52, "ONP (13%)": 44, "AFP Aporte": 44,
        "AFP Seguro": 44, "AFP Comis.": 44, "Ret. 5ta Cat.": 46,
        "Dsctos/Faltas": 46, "NETO A PAGAR": 55, "Aporte Seg. Social": 58,
    }
    col_w = [col_widths_map.get(c, 48) for c in cols_mostrar]
    # Escalar para ocupar exactamente W_PAGE
    total_w = sum(col_w)
    col_w = [w * W_PAGE / total_w for w in col_w]

    # ── Construir datos de la tabla ──
    # Cabeceras más cortas para encabezado (wrapeadas con Paragraph)
    hdr_style = ParagraphStyle('HDR', fontName="Helvetica-Bold", fontSize=6.5,
                               textColor=colors.white, alignment=1, leading=8)
    p_nom_style = ParagraphStyle('NOM', fontName="Helvetica", fontSize=6.5,
                                 textColor=colors.black, alignment=0, leading=8, wordWrap='LTR')
    # Índice de la columna "Apellidos y Nombres" para aplicar word-wrap
    nom_idx = cols_mostrar.index("Apellidos y Nombres") if "Apellidos y Nombres" in cols_mostrar else -1
    data_rows = [[Paragraph(c, hdr_style) for c in cols_mostrar]]

    for _, row in df.iterrows():
        fila = []
        for i, val in enumerate(row):
            if i == nom_idx:
                fila.append(Paragraph(str(val) if str(val) != "nan" else "", p_nom_style))
            elif isinstance(val, float):
                fila.append(f"{val:,.2f}")
            else:
                fila.append(str(val) if str(val) != "nan" else "")
        data_rows.append(fila)

    t = Table(data_rows, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  C_STEEL),
        ('TEXTCOLOR',     (0,0),  (-1,0),  colors.white),
        ('ALIGN',         (0,0),  (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (2,1),  (2,-1),  'LEFT'),    # Nombres a la izquierda
        ('FONTNAME',      (0,1),  (-1,-2), 'Helvetica'),
        ('FONTSIZE',      (0,1),  (-1,-2), 6.5),
        ('TOPPADDING',    (0,0),  (-1,-1), 4),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 4),
        ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, C_LIGHT]),
        ('BACKGROUND',    (0,-1), (-1,-1), colors.HexColor("#CBD5E1")),
        ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,-1), (-1,-1), 6.5),
        ('GRID',          (0,0),  (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ('LINEABOVE',     (0,-1), (-1,-1), 0.8, C_NAVY),
        ('LINEBELOW',     (0,0),  (-1,0),  0.8, C_GOLD),
    ]))
    elements.append(t)
    
    # --- CUADRO RESUMEN DE PREVISIONES (AFP/ONP/SEGURIDAD SOCIAL/5TA) ---
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("<b>RESUMEN GENERAL DE OBLIGACIONES DEL EMPLEADOR Y RETENCIONES</b>", st_sub))
    elements.append(Spacer(1, 10))

    df_data = df.iloc[:-1]  # Sin fila de totales
    resumen_data = [["CONCEPTO", "DETALLE", "MONTO (S/)  "]]
    total_general = 0.0

    # ONP
    if 'ONP (13%)' in df_data.columns:
        onp_total = df_data['ONP (13%)'].sum()
        if onp_total > 0:
            n_onp = len(df_data[df_data['ONP (13%)'] > 0])
            resumen_data.append(["RETENCIÓN ONP (13%)", f"{n_onp} trabajador(es)", f"{onp_total:,.2f}"])
            total_general += onp_total

    # AFP por entidad — columnas pueden no existir si el filtro de ceros las eliminó
    if 'Sist. Pensión' in df_data.columns:
        for afp in df_data['Sist. Pensión'].unique():
            if "AFP" in str(afp):
                df_afp = df_data[df_data['Sist. Pensión'] == afp]
                tot = (
                    (df_afp['AFP Aporte'].sum() if 'AFP Aporte' in df_afp.columns else 0.0) +
                    (df_afp['AFP Seguro'].sum() if 'AFP Seguro' in df_afp.columns else 0.0) +
                    (df_afp['AFP Comis.'].sum() if 'AFP Comis.' in df_afp.columns else 0.0)
                )
                if tot > 0:
                    resumen_data.append([f"RETENCIÓN {afp}", f"{len(df_afp)} trabajador(es)", f"{tot:,.2f}"])
                    total_general += tot

    # Renta 5ta Categoría
    if 'Ret. 5ta Cat.' in df_data.columns:
        quinta_total = df_data['Ret. 5ta Cat.'].sum()
        if quinta_total > 0:
            n_quinta = len(df_data[df_data['Ret. 5ta Cat.'] > 0])
            resumen_data.append(["RETENCIÓN RENTA 5TA CAT.", f"{n_quinta} trabajador(es)", f"{quinta_total:,.2f}"])
            total_general += quinta_total

    resumen_data.append(["TOTAL RETENCIONES (A DECLARAR PDT)", "", f"S/ {total_general:,.2f}"])

    # Seguridad Social (ESSALUD + SIS) — a cargo del empleador
    resumen_data.append(["", "", ""])
    aporte_seg_total = 0.0
    col_seg = 'Aporte Seg. Social' if 'Aporte Seg. Social' in df_data.columns else 'EsSalud Patronal'
    if col_seg in df_data.columns and 'Seg. Social' in df_data.columns:
        for tipo_seg in df_data['Seg. Social'].unique():
            df_seg = df_data[df_data['Seg. Social'] == tipo_seg]
            monto_seg = df_seg[col_seg].sum()
            if monto_seg > 0:
                resumen_data.append([f"APORTE {tipo_seg} (EMPLEADOR)", f"{len(df_seg)} trabajador(es)", f"{monto_seg:,.2f}"])
                aporte_seg_total += monto_seg
    elif col_seg in df_data.columns:
        aporte_seg_total = df_data[col_seg].sum()
        resumen_data.append(["APORTE ESSALUD (EMPLEADOR)", f"{len(df_data)} trabajador(es)", f"{aporte_seg_total:,.2f}"])

    if aporte_seg_total > 0:
        resumen_data.append(["TOTAL SEGURIDAD SOCIAL (A PAGAR SUNAT)", "", f"S/ {aporte_seg_total:,.2f}"])

    t_res = Table(resumen_data, colWidths=[250, 150, 150])
    estilos_res = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#34495E")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#BDC3C7")),
    ]
    # Filas de subtotal en negrita con fondo
    for i, row_data in enumerate(resumen_data):
        if row_data[0].startswith("TOTAL"):
            estilos_res.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#E5E7E9")))
            estilos_res.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
    t_res.setStyle(TableStyle(estilos_res))

    if len(resumen_data) > 2:
        elements.append(t_res)

    doc.build(elements)
    buffer.seek(0)
    return buffer

def generar_pdf_quinta(data_q, empresa_nombre, periodo, nombre_trabajador):
    """Genera el Certificado Corporativo de Retención de 5ta Categoría"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=portrait(letter), rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, textColor=colors.HexColor("#2C3E50"), spaceAfter=5)
    subtitle_style = ParagraphStyle('CustomSub', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor("#7F8C8D"), alignment=TA_CENTER, spaceAfter=20)
    header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=10, textColor=colors.black, spaceAfter=5)

    elements.append(Paragraph(f"<b>{empresa_nombre}</b>", title_style))
    elements.append(Paragraph(f"LIQUIDACIÓN Y DETALLE DE RETENCIÓN DE 5TA CATEGORÍA", subtitle_style))
    elements.append(Paragraph(f"<b>Trabajador:</b> {nombre_trabajador}", header_style))
    elements.append(Paragraph(f"<b>Periodo de Cálculo:</b> {periodo}", header_style))
    elements.append(Spacer(1, 15))

    datos_tabla = [
        ["CONCEPTO / PASO DE CÁLCULO (SUNAT)", "IMPORTES (S/)"],
        ["1. Remuneraciones Previas Percibidas (Ene - Mes Anterior)", f"{int(data_q['rem_previa']):,}"],
        ["2. Remuneración Computable del Mes Actual", f"{data_q['base_mes']:,.2f}"],
        [f"3. Proyección de Meses Restantes ({data_q['meses_restantes']} meses)", f"{data_q['proy_sueldo']:,.2f}"],
        ["4. Proyección de Gratificaciones + Bono Extraordinario 9%", f"{data_q['proy_grati']:,.2f}"],
        ["A. RENTA BRUTA ANUAL PROYECTADA (REDONDEADA)", f"{int(data_q['bruta_anual']):,}"],
        [f"B. Deducción de Ley (7 UIT de S/ {data_q['uit_valor']:,.2f})", f"- {int(data_q['uit_7']):,}"],
        ["C. RENTA NETA IMPONIBLE ANUAL (A - B)", f"{int(data_q['neta_anual']):,}"]
    ]

    t_calc = Table(datos_tabla, colWidths=[380, 100])
    t_calc.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1f77b4")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,5), (-1,5), 'Helvetica-Bold'), 
        ('FONTNAME', (0,7), (-1,7), 'Helvetica-Bold'), 
        ('BACKGROUND', (0,7), (-1,7), colors.HexColor("#ECF0F1")),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7"))
    ]))
    elements.append(t_calc)
    elements.append(Spacer(1, 15))

    if data_q['neta_anual'] > 0:
        elements.append(Paragraph("<b>D. APLICACIÓN DE ESCALAS Y TRAMOS DEL IMPUESTO:</b>", header_style))
        elements.append(Spacer(1, 5))
        
        datos_tramos = [["TRAMO (UIT)", "TASA", "BASE APLICABLE", "IMPUESTO ANUAL"]]
        for tramo in data_q['detalle_tramos']:
            datos_tramos.append([tramo['rango'], tramo['tasa'], f"{int(tramo['base']):,}", f"{int(tramo['impuesto']):,}"])
        
        datos_tramos.append(["", "", "IMPUESTO ANUAL TOTAL:", f"{int(data_q['imp_anual']):,}"])
        
        t_tramos = Table(datos_tramos, colWidths=[150, 60, 130, 140])
        t_tramos.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#7F8C8D")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7"))
        ]))
        elements.append(t_tramos)
        elements.append(Spacer(1, 15))

        datos_final = [
            ["E. Impuesto Anual Calculado", f"{int(data_q['imp_anual']):,}"],
            ["F. Retenciones de Meses Anteriores Efectuadas", f"- {int(data_q['ret_previa']):,}"],
            [f"G. Divisor Aplicable al Mes ({data_q['divisor']})", f"÷ {data_q['divisor']}"],
            ["H. RETENCIÓN EXACTA A EFECTUAR EN EL MES (REDONDEADA)", f"S/ {int(data_q['retencion']):,}"]
        ]
        t_final = Table(datos_final, colWidths=[380, 100])
        t_final.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#F9E79F")), 
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7"))
        ]))
        elements.append(t_final)
    else:
        elements.append(Paragraph("<i>El trabajador no supera las 7 UIT, por lo tanto, la Renta Neta es S/ 0.00 y no aplica retención en este periodo.</i>", header_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer
def generar_excel_honorarios(df_loc, empresa_nombre, periodo_key, empresa_ruc=""):
    """Genera Excel corporativo para valorización de locadores."""
    periodo_texto = _periodo_legible_calc(periodo_key)
    # Limpiar columnas no deseadas para el reporte corporativo
    cols_excluir = ["Banco", "N° Cuenta", "CCI", "Observaciones"]
    df_export = df_loc[[c for c in df_loc.columns if c not in cols_excluir]].copy()
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name=f'Honorarios_{periodo_key[:2]}', index=False, startrow=5)
        ws = writer.sheets[f'Honorarios_{periodo_key[:2]}']
        ws['A1'] = empresa_nombre
        ws['A1'].font = Font(size=16, bold=True, color="0F2744")
        ws['A2'] = f"RUC: {empresa_ruc}" if empresa_ruc else ""
        ws['A2'].font = Font(size=10, color="64748B")
        ws['A3'] = f"VALORIZACIÓN DE LOCADORES DE SERVICIO (4ta Categoría) — PERIODO: {periodo_texto}"
        ws['A3'].font = Font(size=11, bold=True, color="1E4D8C")
        ws['A4'] = f"Fecha de Cálculo: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A4'].font = Font(size=10, italic=True, color="7F8C8D")
        fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_thin
                if cell.row == 6:
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = align_center
                elif cell.row == ws.max_row:
                    cell.fill = fill_total
                    cell.font = Font(bold=True)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 25)
    buffer.seek(0)
    return buffer


def generar_pdf_honorarios(df_loc, empresa_nombre, periodo_key, empresa_ruc="", empresa_regimen=""):
    """Genera PDF corporativo para valorización de locadores (igual diseño que sábana planilla)."""
    periodo_texto = _periodo_legible_calc(periodo_key)
    W_PAGE = 1008 - 24
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=15, bottomMargin=15
    )
    elements = []
    st_title = ParagraphStyle('T', fontName="Helvetica-Bold", fontSize=14, textColor=C_NAVY, spaceAfter=6)
    st_sub   = ParagraphStyle('S', fontName="Helvetica",      fontSize=9,  textColor=C_GRAY, spaceAfter=1)
    st_head  = ParagraphStyle('H', fontName="Helvetica-Bold", fontSize=10, textColor=C_STEEL, spaceAfter=8, spaceBefore=4)
    fecha_calc = datetime.now().strftime("%d/%m/%Y %H:%M")
    ruc_line = f"  |  RUC: {empresa_ruc}" if empresa_ruc else ""
    reg_line = f"  |  {empresa_regimen}" if empresa_regimen else ""
    elements.append(Paragraph(empresa_nombre + ruc_line, st_title))
    elements.append(Paragraph(
        f"VALORIZACIÓN DE LOCADORES DE SERVICIO (4ta CATEGORÍA)  ·  PERIODO: {periodo_texto}{reg_line}", st_head
    ))
    elements.append(Paragraph(f"Fecha de cálculo: {fecha_calc}", st_sub))
    elements.append(Spacer(1, 8))

    # Limpiar columnas para PDF corporativo
    cols_excluir = ["Banco", "N° Cuenta", "CCI", "Observaciones"]
    cols = [c for c in df_loc.columns if c not in cols_excluir]
    
    col_widths_map = {
        "DNI": 55, "Locador": 120,
        "Honorario Base": 65, "Días no Prestados": 55,
        "Días Laborados": 55, "Descuento Días": 60, 
        "Otros Pagos": 60, "Pago Bruto": 60, 
        "Retención 4ta (8%)": 70, "Otros Descuentos": 65, 
        "NETO A PAGAR": 65,
    }
    col_w = [col_widths_map.get(c, 60) for c in cols]
    total_w = sum(col_w)
    col_w = [w * W_PAGE / total_w for w in col_w]

    hdr_style = ParagraphStyle('HDR', fontName="Helvetica-Bold", fontSize=7,
                               textColor=colors.white, alignment=1, leading=8)
    nom_style  = ParagraphStyle('NOM', fontName="Helvetica", fontSize=7,
                                textColor=colors.black, alignment=0, leading=8, wordWrap='LTR')
    nom_idx = cols.index("Locador") if "Locador" in cols else -1
    data_rows = [[Paragraph(c, hdr_style) for c in cols]]

    # Variables de fecha para el cálculo de días reales
    mes_num  = int(periodo_key[:2])
    anio_num = int(periodo_key[3:])
    dias_mes = calendar.monthrange(anio_num, mes_num)[1]

    for _, row in df_loc[cols].iterrows():
        fila = []
        # Días laborados desde el snapshot — dato auditado y congelado al momento del cálculo
        d_lab_real = int(row.get("Días Laborados", 0) or 0)

        for i, c_name in enumerate(cols):
            val = row[c_name]
            if c_name == "Locador":
                fila.append(Paragraph(str(val) if str(val) != "nan" else "", nom_style))
            elif c_name == "Días Laborados":
                fila.append(str(d_lab_real))
            elif isinstance(val, float):
                fila.append(f"{val:,.2f}")
            else:
                fila.append(str(val) if str(val) != "nan" else "")
        data_rows.append(fila)

    # Fila de totales
    tot_style = ParagraphStyle('TC', fontName="Helvetica-Bold", fontSize=7, textColor=colors.white, alignment=1)
    totales_row = []
    cols_texto_loc = {"DNI", "Locador"}
    for c in cols:
        if c == "Locador":
            totales_row.append(Paragraph("<b>TOTALES</b>", tot_style))
        elif c in cols_texto_loc:
            totales_row.append("")
        else:
            try:
                v = df_loc[c].sum()
                totales_row.append(Paragraph(f"<b>{v:,.2f}</b>", tot_style))
            except Exception:
                totales_row.append("")
    data_rows.append(totales_row)

    t = Table(data_rows, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  C_STEEL),
        ('TEXTCOLOR',     (0,0),  (-1,0),  colors.white),
        ('ALIGN',         (0,0),  (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (1,1),  (1,-1),  'LEFT'),
        ('FONTNAME',      (0,1),  (-1,-2), 'Helvetica'),
        ('FONTSIZE',      (0,1),  (-1,-2), 7),
        ('TOPPADDING',    (0,0),  (-1,-1), 4),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 4),
        ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, C_LIGHT]),
        ('BACKGROUND',    (0,-1), (-1,-1), C_NAVY),
        ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,-1), (-1,-1), 7),
        ('GRID',          (0,0),  (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ('LINEABOVE',     (0,-1), (-1,-1), 0.8, C_NAVY),
        ('LINEBELOW',     (0,0),  (-1,0),  0.8, C_GOLD),
    ]))
    elements.append(t)

    # Resumen de retenciones 4ta
    elements.append(Spacer(1, 20))
    st_res = ParagraphStyle('R', fontName="Helvetica-Bold", fontSize=9, textColor=C_NAVY, spaceAfter=6)
    elements.append(Paragraph("RESUMEN DE RETENCIONES — 4ta CATEGORÍA (Ley SUNAT)", st_res))
    try:
        total_bruto = df_loc["Pago Bruto"].sum()
        total_ret   = df_loc["Retención 4ta (8%)"].sum()
        total_neto  = df_loc["NETO A PAGAR"].sum()
        n_loc       = len(df_loc)
        n_con_ret   = len(df_loc[df_loc["Retención 4ta (8%)"] > 0])
        resumen = [
            ["CONCEPTO", "DETALLE", "MONTO (S/)"],
            ["Pago Bruto Total Locadores", f"{n_loc} persona(s)", f"{total_bruto:,.2f}"],
            ["Retención 4ta Categoría (8%)", f"{n_con_ret} locador(es) afecto(s)", f"{total_ret:,.2f}"],
            ["TOTAL NETO A PAGAR", "", f"S/ {total_neto:,.2f}"],
        ]
        t_res = Table(resumen, colWidths=[260, 200, 140])
        estilos_res = [
            ('BACKGROUND', (0,0),  (-1,0),  colors.HexColor("#34495E")),
            ('TEXTCOLOR',  (0,0),  (-1,0),  colors.whitesmoke),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#E5E7E9")),
            ('FONTNAME',   (0,0),  (-1,0),  'Helvetica-Bold'),
            ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0),  (-1,-1), 8),
            ('ALIGN',      (2,0),  (2,-1),  'RIGHT'),
            ('ALIGN',      (0,0),  (1,-1),  'LEFT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('GRID',       (0,0),  (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
        ]
        t_res.setStyle(TableStyle(estilos_res))
        elements.append(t_res)
    except Exception:
        pass

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_pdf_combinado(df_planilla, df_loc, empresa_nombre, periodo_key, empresa_ruc="", empresa_regimen=""):
    """Genera PDF consolidado: Sábana planilla + Valorización locadores + Resumen costo laboral total."""
    periodo_texto = _periodo_legible_calc(periodo_key)
    W_PAGE = 1008 - 24
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=15, bottomMargin=15
    )
    elements = []
    st_title = ParagraphStyle('T',  fontName="Helvetica-Bold", fontSize=13, textColor=C_NAVY, spaceAfter=4)
    st_head  = ParagraphStyle('H',  fontName="Helvetica-Bold", fontSize=10, textColor=C_STEEL, spaceAfter=6)
    st_sec   = ParagraphStyle('SE', fontName="Helvetica-Bold", fontSize=9,  textColor=C_NAVY, spaceAfter=4, spaceBefore=10,
                              borderPad=3, backColor=colors.HexColor("#E8F0FE"))
    st_sub   = ParagraphStyle('S',  fontName="Helvetica",      fontSize=8,  textColor=C_GRAY, spaceAfter=1)

    fecha_calc = datetime.now().strftime("%d/%m/%Y %H:%M")
    ruc_line = f"  |  RUC: {empresa_ruc}" if empresa_ruc else ""
    elements.append(Paragraph(empresa_nombre.upper() + ruc_line, st_title))
    elements.append(Paragraph(
        f"REPORTE CONSOLIDADO DE COSTO LABORAL  ·  PERIODO: {periodo_texto}", st_head
    ))
    elements.append(Paragraph(
        f"Planilla de Remuneraciones (5ta Cat.) + Locadores de Servicio (4ta Cat.)  |  Generado: {fecha_calc}", st_sub
    ))
    elements.append(Spacer(1, 8))

    hdr_s = ParagraphStyle('HS', fontName="Helvetica-Bold", fontSize=6, textColor=colors.white, alignment=1, leading=7)
    nom_s = ParagraphStyle('NS', fontName="Helvetica", fontSize=6, textColor=colors.black, alignment=0, leading=7, wordWrap='LTR')

    # ── SECCIÓN 1: PLANILLA ───────────────────────────────────────────────────
    elements.append(Paragraph("▶  1. PLANILLA DE REMUNERACIONES (5ta Categoría)", st_sec))
    _BANCO_COLS_COMB = _COLS_OCULTAS_SABANA | {"Banco", "N° Cuenta", "CCI", "Observaciones"}
    cols_plan = [c for c in df_planilla.columns if c not in _BANCO_COLS_COMB]
    df_p = df_planilla[cols_plan]
    col_widths_plan = {
        "N°": 18, "DNI": 52, "Apellidos y Nombres": 105,
        "Sist. Pensión": 55, "Seg. Social": 48,
        "Sueldo Base": 50, "Asig. Fam.": 36, "Otros Ingresos": 50,
        "TOTAL BRUTO": 50, "ONP (13%)": 42, "AFP Aporte": 42,
        "AFP Seguro": 42, "AFP Comis.": 42, "Ret. 5ta Cat.": 44,
        "Dsctos/Faltas": 44, "NETO A PAGAR": 52, "Aporte Seg. Social": 55,
    }
    col_w_p = [col_widths_plan.get(c, 46) for c in cols_plan]
    total_w_p = sum(col_w_p)
    col_w_p = [w * W_PAGE / total_w_p for w in col_w_p]
    nom_idx_p = cols_plan.index("Apellidos y Nombres") if "Apellidos y Nombres" in cols_plan else -1
    rows_p = [[Paragraph(c, hdr_s) for c in cols_plan]]
    for _, row in df_p.iterrows():
        fila = []
        for i, val in enumerate(row):
            if i == nom_idx_p:
                fila.append(Paragraph(str(val) if str(val) != "nan" else "", nom_s))
            elif isinstance(val, float):
                fila.append(f"{val:,.2f}")
            else:
                fila.append(str(val) if str(val) != "nan" else "")
        rows_p.append(fila)
    t_p = Table(rows_p, colWidths=col_w_p, repeatRows=1)
    t_p.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  C_STEEL),
        ('ALIGN',         (0,0),  (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (2,1),  (2,-1),  'LEFT'),
        ('FONTNAME',      (0,1),  (-1,-2), 'Helvetica'),
        ('FONTSIZE',      (0,1),  (-1,-2), 6),
        ('TOPPADDING',    (0,0),  (-1,-1), 3),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 3),
        ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, C_LIGHT]),
        ('BACKGROUND',    (0,-1), (-1,-1), colors.HexColor("#CBD5E1")),
        ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,-1), (-1,-1), 6),
        ('GRID',          (0,0),  (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ('LINEBELOW',     (0,0),  (-1,0),  0.8, C_GOLD),
    ]))
    elements.append(t_p)

    # ── SECCIÓN 2: LOCADORES ──────────────────────────────────────────────────
    if df_loc is not None and not df_loc.empty:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("▶  2. VALORIZACIÓN DE LOCADORES DE SERVICIO (4ta Categoría)", st_sec))
        # Excluir columnas bancarias y observaciones del reporte combinado
        _BANCO_COMB = {"Banco", "N° Cuenta", "CCI", "Observaciones"}
        cols_loc = [c for c in df_loc.columns if c not in _BANCO_COMB]
        col_widths_loc = {
            "DNI": 52, "Locador": 120,
            "Honorario Base": 65, "Días no Prestados": 55,
            "Descuento Días": 60, "Otros Pagos": 60,
            "Pago Bruto": 60, "Retención 4ta (8%)": 65,
            "Otros Descuentos": 65, "NETO A PAGAR": 65,
        }
        col_w_l = [col_widths_loc.get(c, 55) for c in cols_loc]
        total_w_l = sum(col_w_l)
        col_w_l = [w * W_PAGE / total_w_l for w in col_w_l]
        nom_idx_l = cols_loc.index("Locador") if "Locador" in cols_loc else -1
        rows_l = [[Paragraph(c, hdr_s) for c in cols_loc]]
        for _, row in df_loc[cols_loc].iterrows():
            fila = []
            for i, val in enumerate(row):
                if i == nom_idx_l:
                    fila.append(Paragraph(str(val) if str(val) != "nan" else "", nom_s))
                elif isinstance(val, float):
                    fila.append(f"{val:,.2f}")
                else:
                    fila.append(str(val) if str(val) != "nan" else "")
            rows_l.append(fila)
        t_l = Table(rows_l, colWidths=col_w_l, repeatRows=1)
        t_l.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),  (-1,0),  C_STEEL),
            ('ALIGN',         (0,0),  (-1,-1), 'CENTER'),
            ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
            ('ALIGN',         (1,1),  (1,-1),  'LEFT'),
            ('FONTNAME',      (0,1),  (-1,-1), 'Helvetica'),
            ('FONTSIZE',      (0,1),  (-1,-1), 6),
            ('TOPPADDING',    (0,0),  (-1,-1), 3),
            ('BOTTOMPADDING', (0,0),  (-1,-1), 3),
            ('ROWBACKGROUNDS',(0,1),  (-1,-1), [colors.white, C_LIGHT]),
            ('GRID',          (0,0),  (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
            ('LINEBELOW',     (0,0),  (-1,0),  0.8, C_GOLD),
        ]))
        elements.append(t_l)

    # ── SECCIÓN 3: RESUMEN COSTO LABORAL TOTAL ────────────────────────────────
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("▶  3. RESUMEN CONSOLIDADO DE COSTO LABORAL", st_sec))
    try:
        df_plan_data = df_planilla[df_planilla.get('Apellidos y Nombres', pd.Series(dtype=str)) != 'TOTALES'] \
            if 'Apellidos y Nombres' in df_planilla.columns else df_planilla.iloc[:-1]
        bruto_plan   = df_plan_data['TOTAL BRUTO'].sum()       if 'TOTAL BRUTO'       in df_plan_data.columns else 0.0
        neto_plan    = df_plan_data['NETO A PAGAR'].sum()      if 'NETO A PAGAR'      in df_plan_data.columns else 0.0
        essalud_plan = df_plan_data['Aporte Seg. Social'].sum() if 'Aporte Seg. Social' in df_plan_data.columns else 0.0
        bruto_loc = df_loc["Pago Bruto"].sum()       if df_loc is not None and not df_loc.empty and "Pago Bruto"       in df_loc.columns else 0.0
        neto_loc  = df_loc["NETO A PAGAR"].sum()     if df_loc is not None and not df_loc.empty and "NETO A PAGAR"     in df_loc.columns else 0.0
        ret_4ta   = df_loc["Retención 4ta (8%)"].sum() if df_loc is not None and not df_loc.empty and "Retención 4ta (8%)" in df_loc.columns else 0.0
        costo_total = bruto_plan + essalud_plan + bruto_loc
        neto_total  = neto_plan + neto_loc

        res_data = [
            ["CONCEPTO", "PLANILLA (5ta Cat.)", "LOCADORES (4ta Cat.)", "TOTAL CONSOLIDADO"],
            ["Masa Salarial / Honorarios Brutos",  f"S/ {bruto_plan:,.2f}",             f"S/ {bruto_loc:,.2f}",  f"S/ {(bruto_plan + bruto_loc):,.2f}"],
            ["Aporte EsSalud (cargo empleador)",    f"S/ {essalud_plan:,.2f}",           "—",                     f"S/ {essalud_plan:,.2f}"],
            ["Retenciones PDT (ONP/AFP/5ta/4ta)",  "Ver sábana planilla",               f"S/ {ret_4ta:,.2f}",    "—"],
            ["Total Neto a Pagar al Personal",      f"S/ {neto_plan:,.2f}",             f"S/ {neto_loc:,.2f}",   f"S/ {neto_total:,.2f}"],
            ["COSTO LABORAL TOTAL (empresa)",       f"S/ {bruto_plan + essalud_plan:,.2f}", f"S/ {bruto_loc:,.2f}", f"S/ {costo_total:,.2f}"],
        ]
        t_res = Table(res_data, colWidths=[200, 145, 145, 165])
        t_res.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),  (-1,0),  colors.HexColor("#34495E")),
            ('TEXTCOLOR',     (0,0),  (-1,0),  colors.whitesmoke),
            ('BACKGROUND',    (0,-1), (-1,-1), C_NAVY),
            ('TEXTCOLOR',     (0,-1), (-1,-1), colors.white),
            ('FONTNAME',      (0,0),  (-1,0),  'Helvetica-Bold'),
            ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0),  (-1,-1), 8),
            ('ALIGN',         (1,0),  (-1,-1), 'RIGHT'),
            ('ALIGN',         (0,0),  (0,-1),  'LEFT'),
            ('BOTTOMPADDING', (0,0),  (-1,-1), 6),
            ('TOPPADDING',    (0,0),  (-1,-1), 5),
            ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, C_LIGHT]),
            ('GRID',          (0,0),  (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
        ]))
        elements.append(t_res)
    except Exception:
        pass

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_pdf_tesoreria(df_planilla, df_loc, empresa_nombre, periodo_key, auditoria_data=None, empresa_ruc=""):
    """Genera PDF de Tesorería (Landscape Legal): planilla con columnas de ingresos dinámicas + locadores con datos bancarios."""
    periodo_texto = _periodo_legible_calc(periodo_key)
    W_PAGE = 1008 - 24

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=95, bottomMargin=15
    )
    elements = []

    st_sub   = ParagraphStyle('TS',  fontName="Helvetica",      fontSize=9,  textColor=C_GRAY, spaceAfter=2)
    st_sec   = ParagraphStyle('TSC', fontName="Helvetica-Bold", fontSize=9,  textColor=C_STEEL, spaceBefore=8, spaceAfter=4)
    hdr_s    = ParagraphStyle('TH',  fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.white,  alignment=TA_CENTER, wordWrap='CJK')
    nom_s    = ParagraphStyle('TN',  fontName="Helvetica",      fontSize=6.5, textColor=colors.black,  wordWrap='CJK')
    tot_s    = ParagraphStyle('TTOT',fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.white,  alignment=TA_CENTER)
    st_obs   = ParagraphStyle('OBS', fontName="Helvetica",      fontSize=7.5, textColor=C_NAVY, leading=11, leftIndent=10, spaceAfter=2)

    _fecha_emision = datetime.now().strftime('%d/%m/%Y %H:%M')
    _dh = _make_canvas_header(
        empresa_nombre, empresa_ruc,
        f"REPORTE DE TESORERÍA — PAGOS DE NÓMINA  |  Periodo: {periodo_texto}",
        _fecha_emision
    )

    # ── TABLA 1: PLANILLA ────────────────────────────────────────────────────────
    elements.append(Paragraph("▶  1. PLANILLA DE REMUNERACIONES (5ta Categoría)", st_sec))

    if df_planilla is not None and not df_planilla.empty:
        df_plan_data = df_planilla[df_planilla.get('Apellidos y Nombres', pd.Series(dtype=str)) != 'TOTALES'] \
            if 'Apellidos y Nombres' in df_planilla.columns else df_planilla.iloc[:-1]

        # Detectar CUALQUIER ingreso que perciba el trabajador de planilla (dinámico total)
        dynamic_income = []
        # 1. Ingresos fijos del DF
        for src, dsp in [("Sueldo Base", "Sueldo Básico"), ("Asig. Fam.", "Asig. Fam.")]:
            if src in df_plan_data.columns and pd.to_numeric(df_plan_data[src], errors='coerce').sum() > 0:
                dynamic_income.append((src, dsp))
        
        # 2. Todos los conceptos dinámicos desde auditoria_data que tengan monto > 0
        all_audit_incomes = set()
        if auditoria_data:
            for dni, info in auditoria_data.items():
                for concepto_nombre, monto in info.get('ingresos', {}).items():
                    if float(monto or 0) > 0 and not concepto_nombre.startswith("Sueldo Base"):
                        all_audit_incomes.add(concepto_nombre)
        
        for inc_name in sorted(list(all_audit_incomes)):
            # Evitar duplicar Asig. Fam si ya se incluyó arriba
            if inc_name != "Asignación Familiar":
                dynamic_income.append((inc_name, inc_name))

        has_5ta   = "Ret. 5ta Cat."  in df_plan_data.columns and df_plan_data["Ret. 5ta Cat."].sum() > 0
        has_dscto = "Dsctos/Faltas"  in df_plan_data.columns and df_plan_data["Dsctos/Faltas"].sum() > 0
        
        # Verificar si hay datos bancarios en planilla
        has_bank_p = df_plan_data["N° Cuenta"].astype(str).str.strip().replace(['nan', 'None', ''], pd.NA).dropna().any() or \
                     df_plan_data["CCI"].astype(str).str.strip().replace(['nan', 'None', ''], pd.NA).dropna().any()

        headers_p = ["N°", "DNI", "Nombres y Apellidos"]
        for _, dsp in dynamic_income:
            headers_p.append(dsp)
        headers_p += ["TOTAL BRUTO", "Ret. Pensiones"]
        if has_5ta:
            headers_p.append("Ret. 5ta Cat.")
        if has_dscto:
            headers_p.append("Otros Dsctos")
        headers_p += ["NETO A PAGAR", "Banco"]
        if has_bank_p:
            headers_p += ["N° Cuenta", "CCI"]

        col_w_map = {
            "N°": 20, "DNI": 48, "Nombres y Apellidos": 110,
            "Sueldo Básico": 58, "Asig. Fam.": 50,
            "Horas Extra 25%": 48, "Horas Extra 35%": 48,
            "Gratificación": 58, "Bono Ext. 9%": 50,
            "TOTAL BRUTO": 62, "Ret. Pensiones": 62,
            "Ret. 5ta Cat.": 52, "Otros Dsctos": 52,
            "NETO A PAGAR": 62, "Banco": 55, "N° Cuenta": 72, "CCI": 82,
        }
        col_w_p = [col_w_map.get(h, 55) for h in headers_p]
        total_wp = sum(col_w_p)
        col_w_p = [w * W_PAGE / total_wp for w in col_w_p]

        rows_p = [[Paragraph(h, hdr_s) for h in headers_p]]
        tot_p = {h: 0.0 for h in headers_p}
        for h in ["N°", "DNI", "Nombres y Apellidos", "Banco", "N° Cuenta", "CCI"]:
            tot_p[h] = None  # texto → no sumar

        for i, (_, row) in enumerate(df_plan_data.iterrows()):
            ret_pens = sum(float(row.get(c, 0.0) or 0.0) for c in ["ONP (13%)", "AFP Aporte", "AFP Seguro", "AFP Comis."])
            fila = [str(i + 1), str(row.get("DNI", "")), Paragraph(str(row.get("Apellidos y Nombres", "")), nom_s)]
            for src, dsp in dynamic_income:
                val = 0.0
                if src in row.index:
                    val = float(row.get(src, 0.0) or 0.0)
                elif auditoria_data:
                    val = float((auditoria_data.get(str(row.get("DNI", "")), {}).get('ingresos', {}) or {}).get(src, 0.0) or 0.0)
                fila.append(f"{val:,.2f}")
                tot_p[dsp] = tot_p.get(dsp, 0.0) + val
            bruto = float(row.get("TOTAL BRUTO", 0.0) or 0.0)
            neto  = float(row.get("NETO A PAGAR", 0.0) or 0.0)
            fila += [f"{bruto:,.2f}", f"{ret_pens:,.2f}"]
            tot_p["TOTAL BRUTO"] = tot_p.get("TOTAL BRUTO", 0.0) + bruto
            tot_p["Ret. Pensiones"] = tot_p.get("Ret. Pensiones", 0.0) + ret_pens
            if has_5ta:
                v5 = float(row.get("Ret. 5ta Cat.", 0.0) or 0.0)
                fila.append(f"{v5:,.2f}")
                tot_p["Ret. 5ta Cat."] = tot_p.get("Ret. 5ta Cat.", 0.0) + v5
            if has_dscto:
                vd = float(row.get("Dsctos/Faltas", 0.0) or 0.0)
                fila.append(f"{vd:,.2f}")
                tot_p["Otros Dsctos"] = tot_p.get("Otros Dsctos", 0.0) + vd
            fila += [f"{neto:,.2f}", str(row.get("Banco", "") or "")]
            if has_bank_p:
                fila += [str(row.get("N° Cuenta", "") or ""), str(row.get("CCI", "") or "")]
            tot_p["NETO A PAGAR"] = tot_p.get("NETO A PAGAR", 0.0) + neto
            rows_p.append(fila)

        # Fila de totales: N° y DNI vacíos, "TOTALES" en columna de nombres (índice 2)
        tot_fila = ["", "", Paragraph("TOTALES", tot_s)]
        for _, dsp in dynamic_income:
            tot_fila.append(f"{tot_p.get(dsp, 0.0):,.2f}")
        tot_fila += [f"{tot_p.get('TOTAL BRUTO', 0.0):,.2f}", f"{tot_p.get('Ret. Pensiones', 0.0):,.2f}"]
        if has_5ta:
            tot_fila.append(f"{tot_p.get('Ret. 5ta Cat.', 0.0):,.2f}")
        if has_dscto:
            tot_fila.append(f"{tot_p.get('Otros Dsctos', 0.0):,.2f}")
        tot_fila += [f"{tot_p.get('NETO A PAGAR', 0.0):,.2f}", ""]
        if has_bank_p:
            tot_fila += ["", ""]
        rows_p.append(tot_fila)

        t1 = Table(rows_p, colWidths=col_w_p, repeatRows=1)
        t1.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0),  (-1, 0),  C_NAVY),
            ('BACKGROUND',    (0, -1), (-1, -1), C_STEEL),
            ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
            ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
            ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
            ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
            ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0),  (-1, -1), 6.5),
            ('ALIGN',         (0, 0),  (-1, -1), 'CENTER'),
            ('ALIGN',         (2, 1),  (2, -2),  'LEFT'),
            ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0),  (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0),  (-1, -1), 3),
            ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, C_LIGHT]),
            ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ('LINEBELOW',     (0, 0),  (-1, 0),  0.8, C_GOLD),
        ]))
        elements.append(t1)
    else:
        elements.append(Paragraph("(Sin datos de planilla para este periodo)", st_sub))

    # ── TABLA 2: LOCADORES ────────────────────────────────────────────────────────
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("▶  2. LOCADORES DE SERVICIO (4ta Categoría)", st_sec))

    if df_loc is not None and not df_loc.empty:
        mes_num  = int(periodo_key[:2])
        anio_num = int(periodo_key[3:])
        dias_mes = calendar.monthrange(anio_num, mes_num)[1]
        has_dias_lab = "Días Laborados" in df_loc.columns
        
        # Verificar si hay datos bancarios en locadores
        has_bank_l = df_loc["N° Cuenta"].astype(str).str.strip().replace(['nan', 'None', ''], pd.NA).dropna().any() or \
                     df_loc["CCI"].astype(str).str.strip().replace(['nan', 'None', ''], pd.NA).dropna().any()

        headers_l = ["N°", "DNI", "Nombres y Apellidos", "Honorario Base"]
        if has_dias_lab:
            headers_l.append("Días Laborados")
        headers_l += ["Otros Pagos", "Pago Bruto", "Retención 4ta", "Otros Dsctos", "NETO A PAGAR", "Banco"]
        if has_bank_l:
            headers_l += ["N° Cuenta", "CCI"]

        col_w_lmap = {
            "N°": 20, "DNI": 48, "Nombres y Apellidos": 110,
            "Honorario Base": 55, "Días Laborados": 45,
            "Otros Pagos": 52,
            "Pago Bruto": 55, "Retención 4ta": 55, "Otros Dsctos": 50, "NETO A PAGAR": 55,
            "Banco": 50, "N° Cuenta": 65, "CCI": 75,
        }
        col_w_l = [col_w_lmap.get(h, 55) for h in headers_l]
        total_wl = sum(col_w_l)
        col_w_l = [w * W_PAGE / total_wl for w in col_w_l]

        loc_col = "Locador" if "Locador" in df_loc.columns else df_loc.columns[2]
        rows_l = [[Paragraph(h, hdr_s) for h in headers_l]]
        tot_hon = tot_otros_p = tot_bruto = tot_ret = tot_dscto = tot_neto = 0.0

        for i, (_, row) in enumerate(df_loc.iterrows()):
            # Días laborados desde el snapshot — dato auditado y congelado al momento del cálculo
            dias_lab = int(row.get("Días Laborados", 0) or 0)
            
            hon_b  = float(row.get("Honorario Base", 0.0) or 0.0)
            otros_p= float(row.get("Otros Pagos", 0.0) or 0.0)
            bruto  = float(row.get("Pago Bruto", 0.0) or 0.0)
            ret    = float(row.get("Retención 4ta (8%)", 0.0) or 0.0)
            dscto  = float(row.get("Otros Descuentos", 0.0) or 0.0)
            neto   = float(row.get("NETO A PAGAR", 0.0) or 0.0)
            fila   = [str(i + 1), str(row.get("DNI", "") or ""),
                      Paragraph(str(row.get(loc_col, "") or ""), nom_s),
                      f"{hon_b:,.2f}"]
            if has_dias_lab:
                fila.append(str(dias_lab))
            fila += [
                f"{otros_p:,.2f}",
                f"{bruto:,.2f}", f"{ret:,.2f}", f"{dscto:,.2f}", f"{neto:,.2f}",
                str(row.get("Banco", "") or ""),
            ]
            if has_bank_l:
                fila += [str(row.get("N° Cuenta", "") or ""), str(row.get("CCI", "") or "")]
            rows_l.append(fila)
            tot_hon += hon_b; tot_otros_p += otros_p; tot_bruto += bruto; tot_ret += ret; tot_dscto += dscto; tot_neto += neto

        # Fila de totales locadores: "TOTALES" en columna de nombres (índice 2)
        tot_l_fila = ["", "", Paragraph("TOTALES", tot_s), f"{tot_hon:,.2f}"]
        if has_dias_lab:
            tot_l_fila.append("")
        tot_l_fila += [f"{tot_otros_p:,.2f}", f"{tot_bruto:,.2f}", f"{tot_ret:,.2f}", f"{tot_dscto:,.2f}", f"{tot_neto:,.2f}", ""]
        if has_bank_l:
            tot_l_fila += ["", ""]
        rows_l.append(tot_l_fila)

        t2 = Table(rows_l, colWidths=col_w_l, repeatRows=1)
        t2.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0),  (-1, 0),  C_NAVY),
            ('BACKGROUND',    (0, -1), (-1, -1), C_STEEL),
            ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
            ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
            ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
            ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
            ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0),  (-1, -1), 6.5),
            ('ALIGN',         (0, 0),  (-1, -1), 'CENTER'),
            ('ALIGN',         (2, 1),  (2, -2),  'LEFT'),
            ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0),  (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0),  (-1, -1), 3),
            ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, C_LIGHT]),
            ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ('LINEBELOW',     (0, 0),  (-1, 0),  0.8, C_GOLD),
        ]))
        elements.append(t2)
    else:
        elements.append(Paragraph("(Sin locadores de servicio para este periodo)", st_sub))

    # ── SECCIÓN 3: OBSERVACIONES DEL PERIODO ─────────────────────────────────────
    obs_all = []
    if auditoria_data:
        for _dni_o, _info_o in auditoria_data.items():
            obs_str = _info_o.get('observaciones', '')
            if obs_str:
                nombre_o = _info_o.get('nombres', str(_dni_o))
                obs_all.append(f"[Planilla] {nombre_o} ({_dni_o}): {obs_str}")
    if df_loc is not None and not df_loc.empty and 'Observaciones' in df_loc.columns:
        _loc_col_o = "Locador" if "Locador" in df_loc.columns else df_loc.columns[2]
        for _, row_o in df_loc.iterrows():
            obs_str = str(row_o.get('Observaciones', '') or '')
            if obs_str:
                obs_all.append(f"[Locador] {str(row_o.get(_loc_col_o, ''))} ({str(row_o.get('DNI', ''))}): {obs_str}")

    if obs_all:
        elements.append(Spacer(1, 14))
        elements.append(Paragraph("▶  3. OBSERVACIONES DEL PERIODO", st_sec))
        for obs_line in obs_all:
            elements.append(Paragraph(f"• {obs_line}", st_obs))

    doc.build(elements, onFirstPage=_dh, onLaterPages=_dh)
    buffer.seek(0)
    return buffer


def generar_pdf_personalizado(df, empresa_nombre, periodo_key, titulo, empresa_ruc=""):
    """
    Genera un PDF corporativo para el Reporte Personalizado con cualquier combinación
    de columnas seleccionadas por el usuario.
    Paleta corporativa C_NAVY / C_STEEL / C_GOLD. Landscape Legal.
    Encabezado corporativo repetido en canvas (topMargin=95).
    Si "Observaciones" está en las columnas, se imprime como sección al final (no en la tabla).
    """
    periodo_texto = _periodo_legible_calc(periodo_key)
    W_PAGE = 1008 - 24  # 984 pt disponibles en landscape legal con márgenes 12

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(legal),
        rightMargin=12, leftMargin=12, topMargin=95, bottomMargin=15
    )
    elements = []

    st_sec   = ParagraphStyle('RPSC', fontName="Helvetica-Bold", fontSize=9,  textColor=C_STEEL, spaceBefore=8, spaceAfter=4)
    hdr_s    = ParagraphStyle('RPH',  fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.white,
                               alignment=TA_CENTER, wordWrap='CJK', leading=8)
    nom_s    = ParagraphStyle('RPN',  fontName="Helvetica",      fontSize=6.5, textColor=colors.black,
                               wordWrap='CJK', leading=8)
    tot_s    = ParagraphStyle('RPTOT',fontName="Helvetica-Bold", fontSize=6.5, textColor=colors.white,
                               alignment=TA_CENTER)
    st_obs   = ParagraphStyle('OBS',  fontName="Helvetica",      fontSize=7.5, textColor=C_NAVY,
                               leading=11, leftIndent=10, spaceAfter=2)

    _fecha_emision = datetime.now().strftime('%d/%m/%Y %H:%M')
    _dh = _make_canvas_header(
        empresa_nombre, empresa_ruc,
        f"{titulo.upper()}  |  Periodo: {periodo_texto}",
        _fecha_emision
    )

    if df is None or df.empty:
        doc.build(elements, onFirstPage=_dh, onLaterPages=_dh)
        buffer.seek(0)
        return buffer

    cols = list(df.columns)

    # "Observaciones" se excluye de la tabla y se renderiza al final como sección
    tiene_obs = "Observaciones" in cols
    cols_tabla = [c for c in cols if c != "Observaciones"]

    # Anchos base por columna (se escalan proporcionalmente a W_PAGE)
    _DEFAULT_W = {
        "N°": 20, "DNI": 48,
        "Apellidos y Nombres": 120, "Locador": 120, "Nombres y Apellidos": 120,
        "Cargo": 75, "Cargo / Actividad": 80,
        "Sist. Pensión": 55, "Seg. Social": 50,
        "Sueldo Base": 52, "Asig. Fam.": 40, "Otros Ingresos": 52,
        "TOTAL BRUTO": 56, "ONP (13%)": 46, "AFP Aporte": 46,
        "AFP Seguro": 46, "AFP Comis.": 46, "Ret. 5ta Cat.": 48,
        "Dsctos/Faltas": 46, "NETO A PAGAR": 58, "Aporte Seg. Social": 58,
        "Honorario Base": 58, "Días no Prestados": 52, "Descuento Días": 52,
        "Otros Pagos": 52, "Pago Bruto": 56, "Retención 4ta (8%)": 56,
        "Otros Descuentos": 56,
        "Banco": 55, "N° Cuenta": 72, "CCI": 82,
    }
    col_w = [_DEFAULT_W.get(c, 58) for c in cols_tabla]
    total_w = sum(col_w)
    col_w = [w * W_PAGE / total_w for w in col_w]

    # Columnas bancarias — se renderizan con Paragraph para word-wrap
    _BANCO_COLS = {"Banco", "N° Cuenta", "CCI"}

    # Columnas de texto (no se suman en la fila de totales)
    _TEXTO = {
        "N°", "DNI", "Apellidos y Nombres", "Locador", "Nombres y Apellidos",
        "Cargo", "Cargo / Actividad", "Sist. Pensión", "Seg. Social",
        "Banco", "N° Cuenta", "CCI",
    }
    # Índice de la columna principal de nombres (para word-wrap y alineación izquierda)
    nom_idx = next(
        (i for i, c in enumerate(cols_tabla) if c in {"Apellidos y Nombres", "Locador", "Nombres y Apellidos"}),
        -1
    )

    # ── Construir filas de la tabla ──────────────────────────────────────────────
    rows_pdf = [[Paragraph(c, hdr_s) for c in cols_tabla]]

    for _, row in df.iterrows():
        fila = []
        for j, c in enumerate(cols_tabla):
            val = row.get(c, "")
            val_str = str(val) if str(val) not in ("nan", "None", "NaN") else ""
            if j == nom_idx or c in _BANCO_COLS:
                fila.append(Paragraph(val_str, nom_s))
            elif isinstance(val, float) and str(val) not in ("nan", "NaN"):
                fila.append(f"{val:,.2f}")
            else:
                fila.append(val_str)
        rows_pdf.append(fila)

    # ── Fila de totales ──────────────────────────────────────────────────────────
    tot_row = []
    for j, c in enumerate(cols_tabla):
        if c in _TEXTO:
            # "TOTALES" va en la columna de nombres; las demás texto quedan vacías
            tot_row.append(Paragraph("TOTALES", tot_s) if j == nom_idx else "")
        else:
            try:
                total = pd.to_numeric(df[c], errors='coerce').fillna(0).sum()
                tot_row.append(f"{total:,.2f}")
            except Exception:
                tot_row.append("")
    # Si no hay columna de nombres, poner TOTALES en índice 2 (o 0 si hay menos columnas)
    if nom_idx < 0 and len(cols_tabla) > 0:
        tot_row[min(2, len(cols_tabla) - 1)] = Paragraph("TOTALES", tot_s)
    rows_pdf.append(tot_row)

    # ── Estilos de la tabla ──────────────────────────────────────────────────────
    style_cmds = [
        ('BACKGROUND',    (0, 0),  (-1, 0),  C_NAVY),
        ('BACKGROUND',    (0, -1), (-1, -1), C_STEEL),
        ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
        ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
        ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1),  (-1, -2), 'Helvetica'),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0),  (-1, -1), 6.5),
        ('ALIGN',         (0, 0),  (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0),  (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0),  (-1, -1), 3),
        ('ROWBACKGROUNDS',(0, 1),  (-1, -2), [colors.white, C_LIGHT]),
        ('GRID',          (0, 0),  (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ('LINEBELOW',     (0, 0),  (-1, 0),  0.8, C_GOLD),
    ]
    if nom_idx >= 0:
        style_cmds.append(('ALIGN', (nom_idx, 1), (nom_idx, -2), 'LEFT'))
    for j, c in enumerate(cols_tabla):
        if c in _BANCO_COLS:
            style_cmds.append(('ALIGN', (j, 1), (j, -2), 'LEFT'))

    t = Table(rows_pdf, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    # ── SECCIÓN: OBSERVACIONES DEL PERIODO ──────────────────────────────────────
    if tiene_obs and 'Observaciones' in df.columns:
        _nom_col = next((c for c in ["Apellidos y Nombres", "Locador", "Nombres y Apellidos"] if c in df.columns), None)
        _dni_col = "DNI" if "DNI" in df.columns else None
        obs_all = []
        for _, row_o in df.iterrows():
            obs_str = str(row_o.get('Observaciones', '') or '').strip()
            if obs_str:
                nombre_o = str(row_o.get(_nom_col, '')) if _nom_col else ''
                dni_o    = str(row_o.get(_dni_col, '')) if _dni_col else ''
                label    = f"{nombre_o} ({dni_o})" if dni_o else nombre_o
                obs_all.append(f"{label}: {obs_str}")
        if obs_all:
            elements.append(Spacer(1, 14))
            elements.append(Paragraph("▶  OBSERVACIONES DEL PERIODO", st_sec))
            for obs_line in obs_all:
                elements.append(Paragraph(f"• {obs_line}", st_obs))

    doc.build(elements, onFirstPage=_dh, onLaterPages=_dh)
    buffer.seek(0)
    return buffer


